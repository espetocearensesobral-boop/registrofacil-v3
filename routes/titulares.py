# registrofacil/routes/titulares.py

from flask import Blueprint, render_template, request, jsonify, session, flash, redirect, url_for
from models import (
    listar_titulares, get_titular_by_id, get_historico_servicos_titular, 
    buscar_titulares_json, executar_query, get_empresa_info, gravar_log,
    titular_tem_processos, editar_titular, excluir_titular
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from flask import Response
from datetime import datetime
from werkzeug.utils import secure_filename
from weasyprint import HTML
import os
from routes.auth import login_status_required, get_client_ip, proteger_input, verificar_csrf_token
from routes.permissoes import permission_required
from utils.logger import operacional_logger as logger
from utils.helpers import get_contrast_color, formatar_data

titulares_bp = Blueprint('titulares', __name__, url_prefix='/titulares')

@titulares_bp.route('/', methods=['GET'])
@login_status_required
@permission_required('titulares_visualizar')
def index():
    logger.info(f"Acessando lista de titulares. Usuário ID: {session.get('usuario_id')}, IP: {get_client_ip()}")
    
    pagina_atual = request.args.get('pagina', 1, type=int)
    registros_por_pagina = request.args.get('registros_por_pagina', 50, type=int)
    busca = proteger_input(request.args.get('busca', ''))
    ordenar = request.args.get('ordenar', 'nome')
    direcao = request.args.get('direcao', 'asc')
    
    filtros = {'busca': busca, 'ordenar': ordenar, 'direcao': direcao}
    
    try:
        resultado = listar_titulares(filtros, pagina_atual, registros_por_pagina)
        titulares = resultado['titulares']
        total_registros = resultado['total_records']
        total_paginas = resultado['total_pages']
        
        return render_template('titulares/index.html',
                               titulares=titulares,
                               total_registros=total_registros,
                               total_paginas=total_paginas,
                               pagina_atual=pagina_atual,
                               registros_por_pagina=registros_por_pagina,
                               busca=busca,
                               ordenar=ordenar,
                               direcao=direcao)
    except Exception as e:
        logger.error(f"Erro ao listar titulares: {e}", exc_info=True)
        flash("Erro ao carregar lista de titulares.", "danger")
        return redirect(url_for('auth.dashboard'))

@titulares_bp.route('/editar/<int:titular_id>', methods=['GET', 'POST'])
@login_status_required
@permission_required('titulares_editar')
def editar(titular_id):
    titular = get_titular_by_id(titular_id)
    if not titular:
        flash("Titular não encontrado.", "warning")
        return redirect(url_for('titulares.index'))

    if request.method == 'POST':
        if not verificar_csrf_token(request.form.get('csrf_token')):
            flash("Token de segurança inválido. Por favor, recarregue o formulário.", "danger")
            logger.warning(f"CSRF inválido ao editar titular {titular_id}. Usuário ID: {session.get('usuario_id')}")
            return render_template('titulares/editar.html', titular=titular)

        expected_updated_at = request.form.get('updated_at', '').strip() or None
        nome = proteger_input(request.form.get('nome', '').strip().upper())
        telefone = proteger_input(request.form.get('telefone', '').strip())
        email = proteger_input(request.form.get('email', '').strip().lower())

        if not nome:
            flash("O nome do titular é obrigatório.", "danger")
            return render_template('titulares/editar.html', titular=titular)

        try:
            from utils.helpers import validar_telefone, validar_email
            from models import validar_telefone_unico, validar_email_unico

            if telefone: validar_telefone(telefone)
            if email: validar_email(email)
            if telefone: validar_telefone_unico(telefone, titular_id=titular_id)
            if email: validar_email_unico(email, titular_id=titular_id)

            # Checar duplicidade de nome (excluindo o próprio)
            existente = executar_query(
                "SELECT id FROM titulares WHERE nome = ? AND id != ?", [nome, titular_id], fetch_one=True
            )
            if existente:
                flash(f"Já existe outro titular com o nome '{nome}'.", "warning")
                return render_template('titulares/editar.html', titular=titular)

            editar_titular(
                titular_id,
                nome,
                telefone if telefone else None,
                email if email else None,
                usuario_id=session.get('usuario_id'),
                expected_updated_at=expected_updated_at,
            )
            gravar_log("Editou titular", None, session.get('usuario_id'), get_client_ip(), f"Titular ID {titular_id}: {nome}")
            flash("Titular atualizado com sucesso!", "success")
            return redirect(url_for('titulares.visualizar', titular_id=titular_id))

        except ValueError as e:
            flash(str(e), "danger")
            return render_template('titulares/editar.html', titular=titular)
        except Exception as e:
            logger.error(f"Erro ao editar titular {titular_id}: {e}", exc_info=True)
            flash("Erro ao atualizar titular.", "danger")

    return render_template('titulares/editar.html', titular=titular)


@titulares_bp.route('/excluir/<int:titular_id>', methods=['POST'])
@login_status_required
@permission_required('titulares_excluir')
def excluir(titular_id):
    from routes.auth import verificar_csrf_token
    if not verificar_csrf_token(request.form.get('csrf_token')):
        flash("Token de segurança inválido.", "danger")
        return redirect(url_for('titulares.index'))

    titular = get_titular_by_id(titular_id)
    if not titular:
        flash("Titular não encontrado.", "warning")
        return redirect(url_for('titulares.index'))

    if titular_tem_processos(titular_id):
        flash("Não é possível excluir um titular com processos vinculados.", "danger")
        return redirect(url_for('titulares.visualizar', titular_id=titular_id))

    try:
        excluir_titular(titular_id)
        gravar_log("Excluiu titular", None, session.get('usuario_id'), get_client_ip(), f"Titular: {titular['nome']}")
        flash(f"Titular '{titular['nome']}' excluído com sucesso!", "success")
    except Exception as e:
        logger.error(f"Erro ao excluir titular {titular_id}: {e}", exc_info=True)
        flash("Erro ao excluir titular.", "danger")

    return redirect(url_for('titulares.index'))


@titulares_bp.route('/visualizar/<int:titular_id>', methods=['GET'])
@login_status_required
@permission_required('titulares_visualizar')
def visualizar(titular_id):
    try:
        titular = get_titular_by_id(titular_id)
        if not titular:
            flash("Titular não encontrado.", "warning")
            return redirect(url_for('titulares.index'))
            
        historico = get_historico_servicos_titular(titular['id'])
        
        return render_template('titulares/visualizar.html',
                               titular=titular,
                               historico=historico,
                               get_contrast_color=get_contrast_color,
                               formatar_data=formatar_data)
    except Exception as e:
        logger.error(f"Erro ao visualizar titular {titular_id}: {e}", exc_info=True)
        flash("Erro ao carregar detalhes do titular.", "danger")
        return redirect(url_for('titulares.index'))

@titulares_bp.route('/novo', methods=['GET', 'POST'])
@login_status_required
@permission_required('titulares_criar')
def novo():
    if request.method == 'POST':
        if not verificar_csrf_token(request.form.get('csrf_token')):
            flash("Token de segurança inválido. Por favor, recarregue o formulário.", "danger")
            logger.warning(f"CSRF inválido ao cadastrar titular. Usuário ID: {session.get('usuario_id')}")
            return render_template('titulares/novo.html')

        nome = proteger_input(request.form.get('nome', '').strip().upper())
        telefone = proteger_input(request.form.get('telefone', '').strip())
        email = proteger_input(request.form.get('email', '').strip().lower())
        
        if not nome:
            flash("O nome do titular é obrigatório.", "danger")
            return render_template('titulares/novo.html')
            
        try:
            from utils.helpers import validar_telefone, validar_email
            from models import validar_telefone_unico, validar_email_unico
            
            # Validações de formato
            if telefone: validar_telefone(telefone)
            if email: validar_email(email)
            
            # Validações de duplicidade
            if telefone: validar_telefone_unico(telefone, titular_nome=nome)
            if email: validar_email_unico(email, titular_nome=nome)
            
            # Verificar se já existe pelo nome
            existente = executar_query("SELECT id FROM titulares WHERE nome = ?", [nome], fetch_one=True)
            if existente:
                flash(f"Já existe um titular cadastrado com o nome '{nome}'.", "warning")
                return render_template('titulares/novo.html')
                
            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            executar_query(
                "INSERT INTO titulares (nome, telefone, email, created_at, updated_at) VALUES (?, ?, ?, ?, ?)",
                [nome, telefone, email, now, now]
            )
            
            gravar_log("Cadastrou novo titular", None, session.get('usuario_id'), get_client_ip(), f"Titular: {nome}")
            flash("Titular cadastrado com sucesso!", "success")
            return redirect(url_for('titulares.index'))
        except Exception as e:
            logger.error(f"Erro ao cadastrar titular: {e}", exc_info=True)
            flash("Erro ao cadastrar titular.", "danger")
            
    return render_template('titulares/novo.html')

@titulares_bp.route('/exportar_excel', methods=['GET'])
@login_status_required
@permission_required('titulares_exportar')
def exportar_excel():
    logger.info(f"Usuário {session.get('usuario_id')} solicitou a exportação de titulares para Excel.")
    busca = proteger_input(request.args.get('busca', ''))
    filtros = {'busca': busca}
    
    try:
        resultado = listar_titulares(filtros, 1, 99999)
        titulares = resultado['titulares']
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Titulares"
        
        headers = ["Nome", "Telefone", "E-mail", "Último Registro (Matrícula)"]
        ws.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="8B6332", end_color="8B6332", fill_type="solid")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            
        for t in titulares:
            ws.append([
                t.get('nome'),
                t.get('telefone') or '-',
                t.get('email') or '-',
                t.get('ultimo_registro_matricula') or '-'
            ])
            
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 20
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"Relatorio_Titulares_{timestamp}.xlsx"
        
        return Response(output,
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment;filename={filename}"})
    except Exception as e:
        logger.error(f"Erro ao exportar titulares: {e}", exc_info=True)
        flash("Erro ao gerar arquivo Excel.", "danger")
        return redirect(url_for('titulares.index'))

@titulares_bp.route('/imprimir', methods=['GET'])
@login_status_required
@permission_required('titulares_imprimir')
def imprimir():
    busca = proteger_input(request.args.get('busca', ''))
    ordenar = request.args.get('ordenar', 'nome')
    direcao = request.args.get('direcao', 'asc')
    filtros = {'busca': busca, 'ordenar': ordenar, 'direcao': direcao}
    
    try:
        resultado = listar_titulares(filtros, 1, 9999)
        titulares = resultado['titulares']
        
        empresa_info = get_empresa_info()
        logo_filename = empresa_info.get('logo') if empresa_info else None
        
        from utils.file_uploads import get_image_url_for_display
        logo_url = get_image_url_for_display(logo_filename, is_company_logo=True)
        
        return render_template('relatorios/titulares_print.html',
                               titulares=titulares,
                               total_registros=len(titulares),
                               logo_url=logo_url,
                               busca=busca,
                               now=datetime.now())
    except Exception as e:
        logger.error(f"Erro ao imprimir titulares: {e}", exc_info=True)
        flash("Erro ao gerar relatório de impressão.", "danger")
        return redirect(url_for('titulares.index'))

@titulares_bp.route('/pdf', methods=['GET'])
@login_status_required
@permission_required('titulares_imprimir')
def gerar_pdf():
    """Gera o PDF da listagem de titulares respeitando os filtros ativos."""
    busca = proteger_input(request.args.get('busca', ''))
    ordenar = request.args.get('ordenar', 'nome')
    direcao = request.args.get('direcao', 'asc')
    filtros = {'busca': busca, 'ordenar': ordenar, 'direcao': direcao}
    try:
        resultado = listar_titulares(filtros, 1, 9999)
        titulares = resultado['titulares']
        empresa_info = get_empresa_info()
        logo_filename = empresa_info.get('logo') if empresa_info else None
        from utils.file_uploads import get_image_url_for_display
        logo_url = get_image_url_for_display(logo_filename, is_company_logo=True)
        html_string = render_template(
            'relatorios/titulares_print.html',
            titulares=titulares,
            total_registros=resultado['total_records'],
            logo_url=logo_url,
            busca=busca,
            now=datetime.now()
        )
        pdf_bytes = HTML(string=html_string, base_url=request.url_root).write_pdf()
        return Response(
            pdf_bytes,
            mimetype='application/pdf',
            headers={'Content-Disposition': f'inline; filename=relatorio_titulares_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'}
        )
    except Exception as e:
        logger.error(f"Erro ao gerar PDF de titulares: {e}", exc_info=True)
        flash("Erro ao gerar relatório em PDF.", "danger")
        return redirect(url_for('titulares.index'))


@titulares_bp.route('/api/buscar', methods=['GET'])
@login_status_required
@permission_required('titulares_visualizar')
def api_buscar():
    termo = request.args.get('q', '')
    # '%' abre listagem completa (até 50); texto normal exige >= 2 chars
    if termo.strip() == '%':
        pass  # allowed
    elif len(termo) < 2:
        return jsonify([])
        
    try:
        resultados = buscar_titulares_json(termo)
        return jsonify(resultados)
    except Exception as e:
        logger.error(f"Erro na API de busca de titulares: {e}")
        return jsonify([]), 500


@titulares_bp.route('/api/verificar-duplicidade', methods=['GET'])
@login_status_required
@permission_required('titulares_visualizar')
def api_verificar_duplicidade():
    """
    Verifica em tempo real se telefone ou e-mail já estão cadastrados,
    excluindo opcionalmente o titular atual (para edição futura).
    """
    campo = request.args.get('campo')  # 'telefone' ou 'email'
    valor = request.args.get('valor', '').strip()
    excluir_id = request.args.get('excluir_id', type=int)

    if campo not in ('telefone', 'email') or not valor:
        return jsonify({'duplicado': False})

    try:
        if campo == 'telefone':
            query = "SELECT id, nome FROM titulares WHERE telefone = ?"
        else:
            query = "SELECT id, nome FROM titulares WHERE email = ?"

        params = [valor]
        if excluir_id:
            query += " AND id != ?"
            params.append(excluir_id)

        resultado = executar_query(query, params, fetch_one=True)
        if resultado:
            return jsonify({'duplicado': True, 'titular': resultado['nome']})
        return jsonify({'duplicado': False})
    except Exception as e:
        logger.error(f"Erro ao verificar duplicidade de {campo}: {e}")
        return jsonify({'duplicado': False})
