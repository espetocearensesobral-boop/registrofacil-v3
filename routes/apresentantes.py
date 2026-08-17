
from flask import Blueprint, render_template, request, jsonify, session, flash, redirect, url_for
from models import (
    listar_apresentantes, get_apresentante_by_id, get_historico_servicos_apresentante, 
    buscar_apresentantes_json, executar_query, get_empresa_info, gravar_log,
    apresentante_tem_processos, editar_apresentante, excluir_apresentante
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from flask import Response
from datetime import datetime
from weasyprint import HTML
from routes.auth import login_status_required, get_client_ip, proteger_input
from routes.permissoes import permission_required
from utils.logger import logger

apresentantes_bp = Blueprint('apresentantes', __name__, url_prefix='/apresentantes')

@apresentantes_bp.route('/', methods=['GET'])
@login_status_required
@permission_required('apresentantes_visualizar')
def index():
    logger.info(f"Acessando lista de apresentantes. Usuário ID: {session.get('usuario_id')}, IP: {get_client_ip()}")
    
    pagina_atual = request.args.get('pagina', 1, type=int)
    registros_por_pagina = request.args.get('registros_por_pagina', 10, type=int)
    busca = proteger_input(request.args.get('busca', ''))
    ordenar = request.args.get('ordenar', 'nome')
    direcao = request.args.get('direcao', 'asc')
    
    filtros = {'busca': busca, 'ordenar': ordenar, 'direcao': direcao}
    
    try:
        resultado = listar_apresentantes(filtros, pagina_atual, registros_por_pagina)
        apresentantes = resultado['apresentantes']
        total_registros = resultado['total_records']
        total_paginas = resultado['total_pages']
        
        return render_template('apresentantes/index.html',
                               apresentantes=apresentantes,
                               total_registros=total_registros,
                               total_paginas=total_paginas,
                               pagina_atual=pagina_atual,
                               registros_por_pagina=registros_por_pagina,
                               busca=busca,
                               ordenar=ordenar,
                               direcao=direcao)
    except Exception as e:
        logger.error(f"Erro ao listar apresentantes: {e}", exc_info=True)
        flash("Erro ao carregar lista de apresentantes.", "danger")
        return redirect(url_for('auth.dashboard'))

@apresentantes_bp.route('/editar/<int:apresentante_id>', methods=['GET', 'POST'])
@login_status_required
@permission_required('apresentantes_editar')
def editar(apresentante_id):
    apresentante = get_apresentante_by_id(apresentante_id)
    if not apresentante:
        flash("Apresentante não encontrado.", "warning")
        return redirect(url_for('apresentantes.index'))

    if request.method == 'POST':
        nome = proteger_input(request.form.get('nome', '').strip().upper())
        telefone = proteger_input(request.form.get('telefone', '').strip())
        email = proteger_input(request.form.get('email', '').strip().lower())

        if not nome:
            flash("O nome do apresentante é obrigatório.", "danger")
            return render_template('apresentantes/editar.html', apresentante=apresentante)

        try:
            from utils.helpers import validar_telefone, validar_email
            if telefone: validar_telefone(telefone)
            if email: validar_email(email)

            existente = executar_query(
                "SELECT id FROM apresentantes WHERE nome = ? AND id != ?", [nome, apresentante_id], fetch_one=True
            )
            if existente:
                flash(f"Já existe outro apresentante com o nome '{nome}'.", "warning")
                return render_template('apresentantes/editar.html', apresentante=apresentante)

            editar_apresentante(
                apresentante_id,
                nome,
                telefone if telefone else None,
                email if email else None,
                usuario_id=session.get('usuario_id'),
            )
            gravar_log("Editou apresentante", None, session.get('usuario_id'), get_client_ip(), f"Apresentante ID {apresentante_id}: {nome}")
            flash("Apresentante atualizado com sucesso!", "success")
            return redirect(url_for('apresentantes.visualizar', apresentante_id=apresentante_id))

        except ValueError as e:
            flash(str(e), "danger")
            return render_template('apresentantes/editar.html', apresentante=apresentante)
        except Exception as e:
            logger.error(f"Erro ao editar apresentante {apresentante_id}: {e}", exc_info=True)
            flash("Erro ao atualizar apresentante.", "danger")

    return render_template('apresentantes/editar.html', apresentante=apresentante)

@apresentantes_bp.route('/excluir/<int:apresentante_id>', methods=['POST'])
@login_status_required
@permission_required('apresentantes_excluir')
def excluir(apresentante_id):
    from routes.auth import verificar_csrf_token
    if not verificar_csrf_token(request.form.get('csrf_token')):
        flash("Token de segurança inválido.", "danger")
        return redirect(url_for('apresentantes.index'))

    apresentante = get_apresentante_by_id(apresentante_id)
    if not apresentante:
        flash("Apresentante não encontrado.", "warning")
        return redirect(url_for('apresentantes.index'))

    if apresentante_tem_processos(apresentante_id):
        flash("Não é possível excluir este apresentante pois existem processos vinculados a ele.", "danger")
        return redirect(url_for('apresentantes.index'))

    try:
        excluir_apresentante(apresentante_id)
        gravar_log("Excluiu apresentante", None, session.get('usuario_id'), get_client_ip(), f"Apresentante excluído: {apresentante['nome']}")
        flash("Apresentante excluído com sucesso!", "success")
    except Exception as e:
        logger.error(f"Erro ao excluir apresentante {apresentante_id}: {e}", exc_info=True)
        flash("Erro ao excluir apresentante.", "danger")

    return redirect(url_for('apresentantes.index'))

@apresentantes_bp.route('/visualizar/<int:apresentante_id>', methods=['GET'])
@login_status_required
@permission_required('apresentantes_visualizar')
def visualizar(apresentante_id):
    apresentante = get_apresentante_by_id(apresentante_id)
    if not apresentante:
        flash("Apresentante não encontrado.", "warning")
        return redirect(url_for('apresentantes.index'))

    from utils.helpers import formatar_data, get_contrast_color
    historico = get_historico_servicos_apresentante(apresentante['id'])
    return render_template('apresentantes/visualizar.html', 
                         apresentante=apresentante, 
                         historico=historico,
                         formatar_data=formatar_data,
                         get_contrast_color=get_contrast_color)

@apresentantes_bp.route('/novo', methods=['GET', 'POST'])
@login_status_required
@permission_required('apresentantes_criar')
def novo():
    if request.method == 'POST':
        nome = proteger_input(request.form.get('nome', '').strip().upper())
        telefone = proteger_input(request.form.get('telefone', '').strip())
        email = proteger_input(request.form.get('email', '').strip().lower())

        if not nome:
            flash("O nome do apresentante é obrigatório.", "danger")
            return render_template('apresentantes/novo.html')

        try:
            from utils.helpers import validar_telefone, validar_email
            if telefone: validar_telefone(telefone)
            if email: validar_email(email)

            existente = executar_query("SELECT id FROM apresentantes WHERE nome = ?", [nome], fetch_one=True)
            if existente:
                flash(f"Já existe um apresentante cadastrado com o nome '{nome}'.", "warning")
                return render_template('apresentantes/novo.html')

            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            with executar_query("INSERT INTO apresentantes (nome, telefone, email, created_at, updated_at) VALUES (?, ?, ?, ?, ?)", 
                                [nome, telefone if telefone else None, email if email else None, now, now], fetch_one=False) as conn:
                pass

            gravar_log("Criou apresentante", None, session.get('usuario_id'), get_client_ip(), f"Novo apresentante: {nome}")
            flash("Apresentante cadastrado com sucesso!", "success")
            return redirect(url_for('apresentantes.index'))

        except ValueError as e:
            flash(str(e), "danger")
            return render_template('apresentantes/novo.html')
        except Exception as e:
            logger.error(f"Erro ao criar apresentante: {e}", exc_info=True)
            flash("Erro ao cadastrar apresentante.", "danger")

    return render_template('apresentantes/novo.html')

@apresentantes_bp.route('/exportar', methods=['GET'])
@login_status_required
@permission_required('apresentantes_exportar')
def exportar():
    try:
        busca = proteger_input(request.args.get('busca', ''))
        ordenar = request.args.get('ordenar', 'nome')
        direcao = request.args.get('direcao', 'asc')
        filtros = {'busca': busca, 'ordenar': ordenar, 'direcao': direcao}
        resultado = listar_apresentantes(filtros, 1, 1000000)
        apresentantes = resultado['apresentantes']

        wb = Workbook()
        ws = wb.active
        ws.title = "Apresentantes"

        headers = ['ID', 'Nome', 'Telefone', 'E-mail', 'Total de Processos', 'Data de Cadastro']
        ws.append(headers)

        header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

        for rep in apresentantes:
            ws.append([
                rep['id'],
                rep['nome'],
                rep['telefone'] or '',
                rep['email'] or '',
                rep['total_processos'],
                rep['created_at']
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"apresentantes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Erro ao exportar apresentantes: {e}", exc_info=True)
        flash("Erro ao exportar apresentantes.", "danger")
        return redirect(url_for('apresentantes.index'))

@apresentantes_bp.route('/imprimir', methods=['GET'])
@login_status_required
@permission_required('apresentantes_imprimir')
def imprimir():
    """Renderiza a listagem completa de apresentantes para impressão."""
    busca = proteger_input(request.args.get('busca', ''))
    ordenar = request.args.get('ordenar', 'nome')
    direcao = request.args.get('direcao', 'asc')
    filtros = {'busca': busca, 'ordenar': ordenar, 'direcao': direcao}

    try:
        resultado = listar_apresentantes(filtros, 1, 9999)
        apresentantes = resultado['apresentantes']

        empresa_info = get_empresa_info()
        logo_filename = empresa_info.get('logo') if empresa_info else None
        from utils.file_uploads import get_image_url_for_display
        logo_url = get_image_url_for_display(logo_filename, is_company_logo=True)

        return render_template(
            'relatorios/apresentantes_print.html',
            apresentantes=apresentantes,
            total_registros=resultado['total_records'],
            logo_url=logo_url,
            busca=busca,
            now=datetime.now()
        )
    except Exception as e:
        logger.error(f"Erro ao gerar impressão de apresentantes: {e}", exc_info=True)
        flash("Erro ao gerar relatório de impressão.", "danger")
        return redirect(url_for('apresentantes.index'))

@apresentantes_bp.route('/pdf', methods=['GET'])
@login_status_required
@permission_required('apresentantes_imprimir')
def gerar_pdf():
    """Gera o PDF da listagem de apresentantes respeitando os filtros ativos."""
    busca = proteger_input(request.args.get('busca', ''))
    ordenar = request.args.get('ordenar', 'nome')
    direcao = request.args.get('direcao', 'asc')
    filtros = {'busca': busca, 'ordenar': ordenar, 'direcao': direcao}
    try:
        resultado = listar_apresentantes(filtros, 1, 9999)
        apresentantes = resultado['apresentantes']
        empresa_info = get_empresa_info()
        logo_filename = empresa_info.get('logo') if empresa_info else None
        from utils.file_uploads import get_image_url_for_display
        logo_url = get_image_url_for_display(logo_filename, is_company_logo=True)
        html_string = render_template(
            'relatorios/apresentantes_print.html',
            apresentantes=apresentantes,
            total_registros=resultado['total_records'],
            logo_url=logo_url,
            busca=busca,
            now=datetime.now()
        )
        pdf_bytes = HTML(string=html_string, base_url=request.url_root).write_pdf()
        return Response(
            pdf_bytes,
            mimetype='application/pdf',
            headers={'Content-Disposition': f'inline; filename=relatorio_apresentantes_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'}
        )
    except Exception as e:
        logger.error(f"Erro ao gerar PDF de apresentantes: {e}", exc_info=True)
        flash("Erro ao gerar relatório em PDF.", "danger")
        return redirect(url_for('apresentantes.index'))


@apresentantes_bp.route('/api/buscar', methods=['GET'])
@login_status_required
def api_buscar():
    termo = request.args.get('q', '')
    apresentantes = buscar_apresentantes_json(termo)
    return jsonify(apresentantes)

@apresentantes_bp.route('/api/verificar-duplicidade', methods=['GET'])
@login_status_required
@permission_required('apresentantes_visualizar')
def api_verificar_duplicidade():
    campo = request.args.get('campo')
    valor = request.args.get('valor', '').strip()
    excluir_id = request.args.get('excluir_id', type=int)

    if campo not in ('telefone', 'email') or not valor:
        return jsonify({'duplicado': False})

    try:
        if campo == 'telefone':
            query = "SELECT id, nome FROM apresentantes WHERE telefone = ?"
        else:
            query = "SELECT id, nome FROM apresentantes WHERE email = ?"

        params = [valor]
        if excluir_id:
            query += " AND id != ?"
            params.append(excluir_id)

        resultado = executar_query(query, params, fetch_one=True)
        if resultado:
            return jsonify({'duplicado': True, 'apresentante': resultado['nome']})
        return jsonify({'duplicado': False})
    except Exception as e:
        logger.error(f"Erro ao verificar duplicidade de apresentante ({campo}): {e}")
        return jsonify({'duplicado': False})
