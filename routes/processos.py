# registrofacil/routes/processos.py

from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_from_directory, current_app, send_file
import functools
import os
import uuid
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import sqlite3
from io import BytesIO
from flask import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from weasyprint import HTML, CSS

# --- Importações para geração de PDF/Word ---
# from xhtml2pdf import pisa # REMOVIDO: A geração de PDF direto foi descontinuada
# REMOVIDO: from docx import Document
# REMOVIDO: from docx.shared import Inches, Pt, RGBColor
# REMOVIDO: from docx.enum.text import WD_ALIGN_PARAGRAPH

# --- Importações do models.py ---
from models import (
    executar_query,
    gravar_log,
    obter_tipos_servico, obter_status_processo_config,
    listar_processos, create_processo, update_processo, get_processo_by_id,
    get_total_processes_count, get_overdue_processes_count, get_in_progress_processes_count,
    get_today_processes_count, get_user_linked_processes_count, get_em_andamento_processes_count,
    obter_usuarios_para_selecao,
    validar_formato_matricula,
    validar_tipo_servico, validar_status,
    get_upload_folder, MAX_FILE_SIZE, ALLOWED_EXTENSIONS,
    upsert_titular_from_processo, upsert_apresentante_from_processo,
    inserir_anexo_processo, obter_anexos_processo, excluir_anexo_processo,
    acquire_lock, release_lock, renew_lock, LOCK_TIMEOUT_MINUTES,
    is_record_locked,
    excluir_processo_db,
    get_sqlite_connection,
    registrar_historico_processo,
    obter_historico_processo,
    get_empresa_info
)

# --- Importações de routes.auth ---
from routes.auth import login_status_required, get_client_ip, proteger_input, verificar_csrf_token, admin_required, gerar_csrf_token

# --- Importações de routes.permissoes ---
from routes.permissoes import permission_required

# --- Importações de utils.logger ---
from utils.logger import logger

# --- Importações de utils.helpers ---
from utils.helpers import (
    get_contrast_color, formatar_data, formatar_tamanho_arquivo, obter_icone_anexo,
    validar_telefone, validar_email,
)

# --- Importações de utils.file_uploads ---
from utils.file_uploads import get_image_url_for_display


processos_bp = Blueprint('processos', __name__, url_prefix='/processos')
ALLOWED_LOCK_TABLES = frozenset({'processos', 'titulares', 'apresentantes'})


def get_mime_type_from_file(filepath, filename):
    """
    Tenta determinar o tipo MIME de um arquivo usando python-magic.
    Faz fallback para inferência por extensão com aviso se python-magic não estiver disponível.
    """
    try:
        import magic
        mime = magic.Magic(mime=True)
        return mime.from_file(filepath)
    except ImportError:
        # Se python-magic não estiver disponível, lançamos um erro crítico em vez de fallback inseguro
        logger.critical("ERRO CRÍTICO DE SEGURANÇA: Biblioteca 'python-magic' não encontrada. Uploads bloqueados por segurança.")
        raise ImportError("Sistema de validação de arquivos indisponível. Contate o suporte.")
    except Exception as e:
        logger.error(f"Erro na detecção de MIME type com python-magic: {e}", exc_info=True)
        raise ValueError('Não foi possível validar o conteúdo do arquivo enviado.') from e

# Mapa de MIME types aceitos por extensão.
# Necessário porque algumas bibliotecas (python-magic) detectam o tipo real do
# contêiner binário em vez do tipo lógico do documento. Por exemplo:
#   - DOCX e XLSX são arquivos ZIP internamente -> magic retorna 'application/zip'
#   - CSV pode ser detectado como 'text/plain' em vez de 'text/csv'
#   - TXT com BOM/encoding pode retornar 'text/plain; charset=utf-8'
MIME_ALIASES = {
    'pdf':  {'application/pdf'},
    'jpg':  {'image/jpeg'},
    'jpeg': {'image/jpeg'},
    'png':  {'image/png'},
}

def mime_valido_para_extensao(extensao, mime_type):
    """Verifica se o MIME type detectado e compativel com a extensao do arquivo."""
    mime_normalizado = mime_type.split(';')[0].strip().lower()  # Remove parametros como charset=utf-8
    mimes_aceitos = MIME_ALIASES.get(extensao, set())
    return bool(mimes_aceitos) and mime_normalizado in mimes_aceitos

def processar_anexos_upload(files_dict, processo_id, usuario_id, conn):
    """Processa o upload de arquivos, salva-os e registra no DB.
    Retorna (anexos_salvos, arquivos_rejeitados).
    O chamador é responsável por registrar logs individuais se necessário."""
    anexos_salvos = []
    arquivos_rejeitados = []
    upload_folder = get_upload_folder()
    os.makedirs(upload_folder, exist_ok=True)
    for file_key in files_dict:
        for uploaded_file in files_dict.getlist(file_key):
            if uploaded_file.filename == '': continue
            
            nome_original = uploaded_file.filename
            
            # Correção: Obter tamanho real do arquivo
            uploaded_file.seek(0, os.SEEK_END)
            tamanho = uploaded_file.tell()
            uploaded_file.seek(0)
            
            extensao = nome_original.rsplit('.', 1)[1].lower() if '.' in nome_original else ''
            
            if extensao not in ALLOWED_EXTENSIONS:
                arquivos_rejeitados.append(f"{nome_original}: Extensão não permitida.")
                logger.warning(f"Upload rejeitado para '{nome_original}': Extensão '{extensao}' não permitida.")
                continue
            
            if tamanho > MAX_FILE_SIZE:
                arquivos_rejeitados.append(f"{nome_original}: Tamanho excede o limite de {MAX_FILE_SIZE / (1024 * 1024):.0f}MB.")
                logger.warning(f"Upload rejeitado para '{nome_original}': Tamanho excessivo ({tamanho} bytes).")
                continue
            
            # Cria um nome de arquivo seguro e único
            nome_arquivo_servidor = str(uuid.uuid4()) + '.' + extensao
            # Salva temporariamente para verificar o MIME type de forma segura
            caminho_completo_temp = os.path.join(get_upload_folder(), secure_filename(nome_arquivo_servidor + ".tmp"))
            
            try:
                uploaded_file.save(caminho_completo_temp)
                
                # Valida o MIME type do arquivo salvo temporariamente
                mime_type = get_mime_type_from_file(caminho_completo_temp, nome_original)

                if not mime_valido_para_extensao(extensao, mime_type):
                    arquivos_rejeitados.append(f"{nome_original}: Tipo de arquivo inválido ou não corresponde à extensão ({mime_type}).")
                    logger.warning(f"Upload rejeitado para '{nome_original}': Tipo MIME incompativel com a extensao '{extensao}' ({mime_type}).")
                    os.remove(caminho_completo_temp)  # Remove o arquivo temporário se a validação falhar
                    continue
                
                # Renomeia o arquivo temporário para o nome final
                caminho_final = os.path.join(get_upload_folder(), secure_filename(nome_arquivo_servidor))
                os.rename(caminho_completo_temp, caminho_final)

                inserir_anexo_processo(
                    processo_id=processo_id,
                    nome_original=nome_original,
                    nome_arquivo_servidor=nome_arquivo_servidor,
                    mime_type=mime_type,
                    tamanho=tamanho,
                    usuario_upload_id=usuario_id,
                    connection=conn
                )
                anexos_salvos.append(nome_original)
                logger.info(f"Anexo '{nome_original}' salvo no DB para processo {processo_id}.")
            except Exception as e:
                arquivos_rejeitados.append(f"{nome_original}: Falha ao validar, salvar no servidor ou registrar no banco.")
                logger.error(f"Erro ao salvar anexo '{nome_original}': {e}", exc_info=True)
                for caminho_residual in (caminho_completo_temp, os.path.join(get_upload_folder(), secure_filename(nome_arquivo_servidor))):
                    if os.path.exists(caminho_residual):
                        try:
                            os.remove(caminho_residual)
                        except Exception as rm_e:
                            logger.error(f"Falha ao remover arquivo residual '{caminho_residual}': {rm_e}")
                continue
    return anexos_salvos, arquivos_rejeitados

def excluir_anexos_selecionados(anexo_ids, processo_id, usuario_id, conn):
    """Exclui anexos selecionados do DB e do sistema de arquivos."""
    anexos_excluidos = []
    erros_exclusao = []
    for anexo_id in anexo_ids:
        try:
            anexo_check = executar_query("SELECT nome_arquivo, nome_original FROM anexos_processos WHERE id = ? AND processo_id = ?", [anexo_id, processo_id], fetch_one=True, connection=conn)
            if anexo_check:
                # Chama a função de exclusão do models que também registra histórico
                nome_arquivo_fisico_removido = excluir_anexo_processo(anexo_id, processo_id, connection=conn)
                if nome_arquivo_fisico_removido:
                    caminho_completo = os.path.join(get_upload_folder(), secure_filename(nome_arquivo_fisico_removido))
                    if os.path.exists(caminho_completo):
                        os.remove(caminho_completo)
                        logger.info(f"Anexo físico '{nome_arquivo_fisico_removido}' removido do servidor.")
                    else:
                        logger.warning(f"Anexo físico '{nome_arquivo_fisico_removido}' não encontrado no servidor, mas removido do DB.")
                    anexos_excluidos.append(anexo_check['nome_original'])
                else:
                    erros_exclusao.append(f"Anexo ID {anexo_id} não pôde ser excluído do DB ou não pertence ao processo {processo_id}.")
                    logger.warning(f"Falha ao excluir anexo do DB: ID {anexo_id}, Processo {processo_id}")
            else:
                erros_exclusao.append(f"Anexo ID {anexo_id} não encontrado ou não pertence ao processo {processo_id}.")
                logger.warning(f"Tentativa de excluir anexo inválido ou não pertencente: ID {anexo_id}, Processo {processo_id}")
        except Exception as e:
            erros_exclusao.append(f"Erro ao excluir anexo ID {anexo_id}: {e}")
            logger.error(f"Erro ao excluir anexo ID {anexo_id} do processo {processo_id}: {e}", exc_info=True)
    return anexos_excluidos, erros_exclusao


@processos_bp.route('/', methods=['GET'])
@login_status_required
@permission_required('processos_visualizar')
def todos():
    logger.info(f"Acessando lista de todos os processos. Usuário ID: {session.get('usuario_id')}, IP: {get_client_ip()}.")
    pagina_atual = request.args.get('pagina', 1, type=int)
    registros_por_pagina = request.args.get('registros_por_pagina', 10, type=int)
    if not (1 <= registros_por_pagina <= 100):
        registros_por_pagina = 10
        flash("Lote de registros inválido; usando o lote padrão de 10.", 'warning')
    
    status_filter = request.args.get('status', '').strip()
    filtro_status_id = None
    if status_filter:
        try:
            filtro_status_id = int(status_filter)
        except ValueError:
            flash("ID de status inválido.", 'danger')
            logger.warning(f"Tentativa de filtrar por status com ID inválido: {status_filter}")
            return redirect(url_for('processos.todos'))

    status_ids_in_str = request.args.get('status_ids_in', '').strip()
    status_ids_in = []
    if status_ids_in_str:
        try:
            status_ids_in = [int(s_id) for s_id in status_ids_in_str.split(',') if s_id.isdigit()]
        except ValueError:
            flash("Filtro de status inválido.", 'danger')
            status_ids_in = []

    filtro_tipo = request.args.get('tipo', type=int)
    filtro_busca = proteger_input(request.args.get('busca'))
    
    # --- Validação e formatação de datas de filtro ---
    filtro_data_inicio = request.args.get('data_inicio')
    if filtro_data_inicio:
        try:
            datetime.strptime(filtro_data_inicio, '%Y-%m-%d')
        except ValueError:
            flash("Formato de 'Data Início' inválido. Use AAAA-MM-DD.", 'danger')
            filtro_data_inicio = None

    filtro_data_fim = request.args.get('data_fim')
    if filtro_data_fim:
        try:
            datetime.strptime(filtro_data_fim, '%Y-%m-%d')
        except ValueError:
            flash("Formato de 'Data Fim' inválido. Use AAAA-MM-DD.", 'danger')
            filtro_data_fim = None
    
    if filtro_data_inicio and filtro_data_fim:
        if datetime.strptime(filtro_data_inicio, '%Y-%m-%d') > datetime.strptime(filtro_data_fim, '%Y-%m-%d'):
            flash("Data de início não pode ser posterior à data de fim.", 'danger')
            filtro_data_inicio = None
            filtro_data_fim = None
    # --- Fim Validação de Datas ---

    filtro_envolve_notas = request.args.get('envolve_notas')
    if filtro_envolve_notas is not None:
        try:
            filtro_envolve_notas = int(filtro_envolve_notas)
            if filtro_envolve_notas not in [0, 1]:
                filtro_envolve_notas = None
        except ValueError:
            filtro_envolve_notas = None
    ordenar = request.args.get('ordenar', 'id_desc')
    ordenar_opcoes = ['data_entrada_asc', 'data_entrada_desc', 'titular_asc', 'titular_desc',
                      'tipo_asc', 'tipo_desc', 'status_asc', 'status_desc', 'id_asc', 'id_desc',
                      'matricula_asc', 'matricula_desc', 'prazo_asc', 'prazo_desc']
    if ordenar not in ordenar_opcoes: ordenar = 'id_desc'
    
    # Este filtro é específico do dashboard e deve ser '1' para ativá-lo
    filtro_pendentes_dashboard = request.args.get('filtro_pendentes_dashboard', type=int)
    
    filters = {
        'tipo': filtro_tipo, 'busca': filtro_busca,
        'data_inicio': filtro_data_inicio, 'data_fim': filtro_data_fim,
        'envolve_notas': filtro_envolve_notas, 'filtro_pendentes_dashboard': filtro_pendentes_dashboard,
    }
    
    # Lógica de prioridade: Se status_ids_in estiver presente, use-o e ignore status_id único
    if status_ids_in:
        filters['status_ids_in'] = status_ids_in
        logger.debug(f"Filtrando processos por status_ids_in (prioridade): {status_ids_in}")
        # Se status_ids_in está ativo, o filtro_status_id do dropdown não deve ser considerado
        filtro_status_id = None 
    else:
        filters['status_id'] = filtro_status_id # Usa o status do dropdown
    
    try:
        resultado = listar_processos(filters, pagina_atual, registros_por_pagina, ordenar)
        processos = resultado['processos'] 
        total_registros = resultado['total_records']
        total_paginas = resultado['total_pages']
        tipos_servico = obter_tipos_servico()
        status_list = obter_status_processo_config()
        total_processos = get_total_processes_count()
        prazo_vencido = get_overdue_processes_count()
        processos_andamento = get_in_progress_processes_count()
        
        # Determinar se filtros ativos estão presentes
        has_active_filters = any([
            status_filter and not status_ids_in,
            bool(status_ids_in),
            filtro_tipo, filtro_busca, filtro_data_inicio,
            filtro_data_fim, filtro_envolve_notas is not None,
            filtro_pendentes_dashboard,
            ordenar != 'id_desc',
            registros_por_pagina != 10
        ])

        acao = request.args.get('acao')
        if acao == 'cadastrado': flash('Processo cadastrado com sucesso.', 'success')
        elif acao == 'atualizado': flash('Registro atualizado com sucesso.', 'success')
        elif acao == 'excluido': flash('Registro excluído com sucesso.', 'success')
        elif acao == 'exportado_excel': flash('Dados exportados para Excel com sucesso!', 'success')
        
        if not processos and has_active_filters:
            flash('Nenhum processo encontrado com os filtros aplicados.', 'info')
    except Exception as e:
        logger.exception(f"Erro ao carregar processos: {e}")
        flash('Erro ao carregar processos. Por favor, tente novamente ou contate o suporte.', 'danger')
        processos = []
        total_registros = 0
        total_paginas = 0
        tipos_servico = obter_tipos_servico()
        status_list = obter_status_processo_config()
        total_processos = 0
        prazo_vencido = 0
        processos_andamento = 0
        has_active_filters = False


    return render_template('processos/todos.html',
                           processos=processos, total_registros=total_registros, total_paginas=total_paginas,
                           pagina_atual=pagina_atual, registros_por_pagina=registros_por_pagina,
                           filtro_status_id=filtro_status_id,
                           status_filter=status_filter,
                           filtro_tipo=filtro_tipo, filtro_busca=filtro_busca,
                           filtro_data_inicio=filtro_data_inicio, filtro_data_fim=filtro_data_fim,
                           filtro_envolve_notas=filtro_envolve_notas, ordenar=ordenar,
                           has_active_filters=has_active_filters,
                           tipos_servico=tipos_servico,
                           status_list=status_list,
                           total_processos=total_processos,
                           prazo_vencido=prazo_vencido, processos_andamento=processos_andamento,
                           get_contrast_color=get_contrast_color, formatar_data=formatar_data,
                           filtro_pendentes_dashboard=filtro_pendentes_dashboard,
                           status_ids_in_active=status_ids_in_str
                           )

@processos_bp.route('/novo', methods=['GET', 'POST'])
@login_status_required
@permission_required('processos_criar')
def novo():
    usuario_id = session.get('usuario_id')
    is_ajax_request = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    tipos_processo = [t for t in obter_tipos_servico() if t.get('ativo')]
    status_processo = [s for s in obter_status_processo_config() if s.get('ativo')]
    responsaveis = obter_usuarios_para_selecao()

    dados_form = {
        'titular': '', 'tipo_servico': '', 'matricula': '',
        'possui_matricula': '0',
            'titular_telefone': '', 'titular_email': '',
        'status': '', 'apresentante': '', 'apresentante_telefone': '', 'apresentante_email': '',
        'prazo_final': '', 'envolvido_notas': 0, 'observacoes': '',
        'data_entrada': datetime.now().strftime('%Y-%m-%d')
    }

    if request.method == 'POST':
        dados_form.update({
            'titular': request.form.get('titular', ''),
            'titular_id': request.form.get('titular_id', ''),
            'titular_telefone': request.form.get('titular_telefone', ''),
            'titular_email': request.form.get('titular_email', ''),
            'tipo_servico': request.form.get('tipo_servico', type=int),
            'matricula': request.form.get('matricula', ''),
            'status': request.form.get('status', ''),
            'apresentante': request.form.get('apresentante', ''),
            'apresentante_id': request.form.get('apresentante_id', ''),
            'apresentante_telefone': request.form.get('apresentante_telefone', ''),
            'apresentante_email': request.form.get('apresentante_email', ''),
            'prazo_final': request.form.get('prazo_final', ''),
            'envolvido_notas': request.form.get('envolvido_notas', type=int),
            'observacoes': request.form.get('observacoes', ''),
            'data_entrada': request.form.get('data_entrada', datetime.now().strftime('%Y-%m-%d'))
        })

        if not verificar_csrf_token(request.form.get('csrf_token')):
            logger.error(f"CSRF Token inválido na criação de novo processo. Usuário ID: {session.get('usuario_id')}, IP: {get_client_ip()}")
            if is_ajax_request:
                return jsonify(success=False, message="Token de segurança inválido. Por favor, recarregue a página e tente novamente.", type='danger'), 403
            else:
                flash("Token de segurança inválido.", 'danger')
                return render_template('processos/novo.html', dados_form=dados_form, tipos_processo=tipos_processo, status_processo=status_processo, responsaveis=responsaveis, csrf_token=gerar_csrf_token())
        
        titular = proteger_input(dados_form['titular'])
        titular_telefone = proteger_input(dados_form['titular_telefone'])
        titular_email = proteger_input(dados_form['titular_email'])
        tipo_id = dados_form['tipo_servico']
        possui_matricula = 1 if request.form.get('possui_matricula') == '1' else 0
        matricula = proteger_input(dados_form['matricula']) if possui_matricula else None
        status_nome = proteger_input(dados_form['status'])
        apresentante = proteger_input(dados_form['apresentante'])
        apresentante_telefone = proteger_input(dados_form['apresentante_telefone'])
        apresentante_email = proteger_input(dados_form['apresentante_email'])
        prazo_final_input = dados_form['prazo_final']
        envolvido_notas = dados_form['envolvido_notas']
        observacoes = proteger_input(dados_form['observacoes'])
        data_entrada = dados_form['data_entrada']

        campos_obrigatorios = {
            'titular': 'Titular', 'tipo_servico': 'Tipo de Serviço',
            'status': 'Status',
            'data_entrada': 'Data de Entrada'
        }
        campos_faltantes = [nome for campo, nome in campos_obrigatorios.items() if not dados_form.get(campo) and dados_form.get(campo) != 0]
        if campos_faltantes:
            msg = f"Os seguintes campos são obrigatórios: {', '.join(campos_faltantes)}"
            if is_ajax_request:
                return jsonify(success=False, message=msg, type='danger', field_error='form_fields_missing'), 400
            else:
                flash(msg, 'danger')
                return render_template('processos/novo.html', dados_form=dados_form, tipos_processo=tipos_processo, status_processo=status_processo, responsaveis=responsaveis, csrf_token=gerar_csrf_token())
        
        try:
            validar_tipo_servico(tipo_id)
            validar_status(status_nome)
            from models import validar_telefone_unico, validar_email_unico
            if matricula:
                validar_formato_matricula(matricula)
            if apresentante_telefone: 
                validar_telefone(apresentante_telefone)
                validar_telefone_unico(apresentante_telefone, titular_nome=titular)
            if apresentante_email: 
                validar_email(apresentante_email)
                validar_email_unico(apresentante_email, titular_nome=titular)
            
            if not data_entrada:
                raise ValueError("A Data de Entrada é obrigatória.")
            try:
                data_entrada_obj = datetime.strptime(data_entrada, '%Y-%m-%d')
            except ValueError:
                raise ValueError("Data de Entrada inválida. Use o formato AAAA-MM-DD.")

            prazo_final_calculado = None
            if prazo_final_input:
                try:
                    data_prazo_obj = datetime.strptime(prazo_final_input, '%Y-%m-%d')
                    prazo_final_calculado = data_prazo_obj.strftime('%Y-%m-%d')
                except ValueError as e:
                    raise ValueError(f"Formato de Prazo Final inválido. Use AAAA-MM-DD. Erro: {e}")
            else:
                tipo_servico_info = executar_query("SELECT prazo_padrao FROM tipos_servico WHERE id = ?", [tipo_id], fetch_one=True)
                dias_prazo = tipo_servico_info['prazo_padrao'] if tipo_servico_info else 30
                prazo_final_calculado = (data_entrada_obj + timedelta(days=dias_prazo)).strftime('%Y-%m-%d')
                logger.info(f"Prazo final calculado automaticamente para {prazo_final_calculado} (Data Entrada: {data_entrada}, Prazo Padrão: {dias_prazo} dias).")
        
        except ValueError as e:
            field_error_map = {
                "Telefone inválido": "apresentante_telefone", "E-mail inválido": "apresentante_email",
                "MATRÍCULA:": "matricula", "TELEFONE:": "apresentante_telefone", "E-MAIL:": "apresentante_email",
                "Tipo de serviço com ID": "tipo_servico",
                "Status": "status",
                "Formato de Prazo Final inválido.": "prazo_final", "O Prazo Final é obrigatório.": "prazo_final",
                "A Data de Entrada é obrigatória.": "data_entrada", "Data de Entrada inválida.": "data_entrada",
                "Matrícula inválida": "matricula"
            }
            field = next((v for k, v in field_error_map.items() if k in str(e)), None)
            if is_ajax_request:
                return jsonify(success=False, message=str(e), type='danger', field_error=field), 400
            else:
                flash(str(e), 'danger')
                return render_template('processos/novo.html', dados_form=dados_form, tipos_processo=tipos_processo, status_processo=status_processo, responsaveis=responsaveis, csrf_token=gerar_csrf_token())
        
        status_id_para_inserir = executar_query("SELECT id FROM status_processo WHERE nome = ?", [status_nome], fetch_one=True)['id']
        data_conclusao = None
        if status_nome.lower() == 'finalizado':
            data_conclusao = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        try:
            with get_sqlite_connection() as conn:
                numero_processo_gerado = str(uuid.uuid4())
                
                processo_id = create_processo(
                    numero_processo=numero_processo_gerado,
                    titular=titular,
                    titular_telefone=titular_telefone,
                    titular_email=titular_email,
                    matricula=matricula,
                    tipo_id=tipo_id,
                    data_entrada=data_entrada,
                    status_id=status_id_para_inserir,
                    prazo_final=prazo_final_calculado,
                    apresentante=apresentante,
                    apresentante_telefone=apresentante_telefone,
                    apresentante_email=apresentante_email,
                    responsavel_id=usuario_id,
                    envolvido_notas=envolvido_notas,
                    observacoes=observacoes,
                    data_conclusao=data_conclusao,
                    possui_matricula=possui_matricula,
                    connection=conn
                )
                
                if not processo_id: 
                    raise ValueError("Falha ao criar processo no banco de dados.")
                
                # Os textos continuam sendo mantidos para compatibilidade, mas o vínculo
                # oficial passa a ser preservado pelos IDs dos cadastros.
                titular_id = upsert_titular_from_processo(
                    titular, titular_telefone, titular_email, processo_id, connection=conn
                )
                apresentante_id = upsert_apresentante_from_processo(
                    apresentante, apresentante_telefone, apresentante_email, processo_id, connection=conn
                )
                executar_query(
                    "UPDATE processos SET titular_id = ?, apresentante_id = ? WHERE id = ?",
                    [titular_id, apresentante_id, processo_id],
                    connection=conn,
                )
                
                anexos_salvos, arquivos_rejeitados = processar_anexos_upload(request.files, processo_id, usuario_id, conn)
                
                redirect_url = url_for('processos.visualizar', processo_id=processo_id)
                logger.debug(f"Generated redirect URL: {redirect_url}")

            # Logs gravados APÓS o commit da transação (conexão própria, não interfere nos dados)
            logger.info(f"Processo '{titular}' (ID: {processo_id}) cadastrado.")
            _ctx_cadastro = None
            if anexos_salvos:
                _n = len(anexos_salvos)
                _ctx_cadastro = f"[anexos] Sim ({_n} arquivo{'s' if _n != 1 else ''})"
            mat_log = matricula if matricula else 'Sem matrícula'
            gravar_log("Cadastrou processo", processo_id, usuario_id, get_client_ip(), f"Processo: {titular} (Matrícula: {mat_log})", contexto=_ctx_cadastro)

            msg_sucesso = "Processo cadastrado com sucesso!"
            if anexos_salvos: msg_sucesso += f" {len(anexos_salvos)} anexo(s) incluído(s)."
            if arquivos_rejeitados:
                msg_sucesso += f" Observação: {len(arquivos_rejeitados)} arquivo(s) não foram anexados. Detalhes: {', '.join(arquivos_rejeitados)}."
            
            return jsonify(success=True, message=msg_sucesso, type='success', redirect=redirect_url)
        
        except sqlite3.IntegrityError as e:
            logger.warning(f"Tentativa de criar processo com número '{numero_processo_gerado}' já existente. Erro: {e}")
            error_msg = "Erro ao cadastrar processo: erro de integridade."
            if is_ajax_request:
                return jsonify(success=False, message=error_msg, type='danger', field_error=None), 400
            else:
                flash(error_msg, 'danger')
                return render_template('processos/novo.html', dados_form=dados_form, tipos_processo=tipos_processo, status_processo=status_processo, responsaveis=responsaveis, csrf_token=gerar_csrf_token())
        except Exception as e:
            error_msg = f"Erro ao cadastrar processo: {str(e)}"
            logger.error(error_msg, exc_info=True)
            if is_ajax_request:
                return jsonify(success=False, message=error_msg, type='danger'), 500
            else:
                flash(error_msg, 'danger')
                return render_template('processos/novo.html', dados_form=dados_form, tipos_processo=tipos_processo, status_processo=status_processo, responsaveis=responsaveis, csrf_token=gerar_csrf_token())


    csrf_token_val = gerar_csrf_token()
    return render_template('processos/novo.html',
                           tipos_processo=tipos_processo, status_processo=status_processo,
                           responsaveis=responsaveis, dados_form=dados_form, csrf_token=csrf_token_val)

@processos_bp.route('/hoje', methods=['GET'])
@login_status_required
@permission_required('processos_visualizar')
def hoje():
    logger.info(f"Acessando processos de hoje. Usuário ID: {session.get('usuario_id')}, IP: {get_client_ip()}")
    pagina_atual = request.args.get('pagina', 1, type=int)
    registros_por_pagina = request.args.get('registros_por_pagina', 50, type=int)
    if not (1 <= registros_por_pagina <= 100):
        registros_por_pagina = 50
        flash("Lote de registros inválido; usando o lote padrão de 50.", 'warning')
    filtro_status_id = request.args.get('status', type=int)
    filtro_tipo = request.args.get('tipo', type=int)
    filtro_busca = proteger_input(request.args.get('busca'))
    
    hoje_formatado = datetime.now().strftime('%Y-%m-%d')

    # --- Validação e formatação de datas de filtro ---
    filtro_data_inicio = request.args.get('data_inicio', hoje_formatado)
    if filtro_data_inicio:
        try:
            datetime.strptime(filtro_data_inicio, '%Y-%m-%d')
        except ValueError:
            flash("Formato de 'Data Início' inválido. Use AAAA-MM-DD.", 'danger')
            filtro_data_inicio = hoje_formatado

    filtro_data_fim = request.args.get('data_fim', hoje_formatado)
    if filtro_data_fim:
        try:
            datetime.strptime(filtro_data_fim, '%Y-%m-%d')
        except ValueError:
            flash("Formato de 'Data Fim' inválido. Use AAAA-MM-DD.", 'danger')
            filtro_data_fim = hoje_formatado
    
    if filtro_data_inicio and filtro_data_fim:
        if datetime.strptime(filtro_data_inicio, '%Y-%m-%d') > datetime.strptime(filtro_data_fim, '%Y-%m-%d'):
            flash("Data de início não pode ser posterior à data de fim.", 'danger')
            filtro_data_inicio = hoje_formatado
            filtro_data_fim = hoje_formatado
    # --- Fim Validação de Datas ---

    filtro_envolve_notas = request.args.get('envolve_notas')
    if filtro_envolve_notas is not None:
        try:
            filtro_envolve_notas = int(filtro_envolve_notas)
            if filtro_envolve_notas not in [0, 1]:
                filtro_envolve_notas = None
        except ValueError:
            filtro_envolve_notas = None
    ordenar = request.args.get('ordenar', 'prazo_asc')
    ordenar_opcoes = ['data_entrada_asc', 'data_entrada_desc', 'titular_asc', 'titular_desc', 'tipo_asc', 'tipo_desc', 'status_asc', 'status_desc', 'id_asc', 'id_desc', 'matricula_asc', 'matricula_desc', 'prazo_asc', 'prazo_desc']
    if ordenar not in ordenar_opcoes: ordenar = 'prazo_asc'
    
    filters = {
        'status_id': filtro_status_id, 'tipo': filtro_tipo, 'busca': filtro_busca,
        'data_inicio': filtro_data_inicio, 'data_fim': filtro_data_fim, 'envolve_notas': filtro_envolve_notas,
    }
    
    try:
        resultado = listar_processos(filters, pagina_atual, registros_por_pagina, ordenar)
        processos = resultado['processos']
        total_registros = resultado['total_records']
        total_paginas = resultado['total_pages']
        tipos_servico = obter_tipos_servico()
        status_list = obter_status_processo_config()
        prazo_vencido = get_overdue_processes_count()
        processos_andamento = get_in_progress_processes_count()
        processos_criados_hoje = get_today_processes_count() 

        has_active_filters = any([
            filtro_status_id, filtro_tipo, filtro_busca, filtro_envolve_notas is not None,
            ordenar != 'prazo_asc', registros_por_pagina != 50,
            filtro_data_inicio != hoje_formatado, filtro_data_fim != hoje_formatado
        ])
        
        acao = request.args.get('acao')
        if acao == 'cadastrado': flash('Processo cadastrado com sucesso.', 'success')
        elif acao == 'atualizado': flash('Registro atualizado com sucesso.', 'success')
        elif acao == 'excluido': flash('Registro excluído com sucesso.', 'success')
        elif acao == 'exportado_excel': flash('Dados exportados para Excel com sucesso!', 'success')
        if not processos and has_active_filters:
            flash('Nenhum processo com prazo para hoje encontrado com os filtros aplicados.', 'info')
    except Exception as e:
        logger.exception(f"Erro ao carregar processos de hoje: {e}")
        flash('Erro ao carregar processos. Por favor, tente novamente ou contate o suporte.', 'danger')
        return redirect(url_for('auth.dashboard'))
    return render_template('processos/hoje.html',
                           processos=processos, total_registros=total_registros, total_paginas=total_paginas,
                           pagina_atual=pagina_atual, registros_por_pagina=registros_por_pagina,
                           filtro_status_id=filtro_status_id, filtro_tipo=filtro_tipo, filtro_busca=filtro_busca,
                           filtro_data_inicio=filtro_data_inicio, filtro_data_fim=filtro_data_fim,
                           filtro_envolve_notas=filtro_envolve_notas, ordenar=ordenar,
                           has_active_filters=has_active_filters,
                           tipos_servico=tipos_servico,
                           status_list=status_list,
                           prazo_vencido=prazo_vencido, processos_andamento=processos_andamento,
                           processos_criados_hoje=processos_criados_hoje,
                           get_contrast_color=get_contrast_color, formatar_data=formatar_data,
                           hoje_formatado=hoje_formatado)

@processos_bp.route('/pendentes', methods=['GET'])
@login_status_required
@permission_required('processos_visualizar')
def pendentes():
    logger.info(f"Acessando processos pendentes. Usuário ID: {session.get('usuario_id')}, IP: {get_client_ip()}")
    pagina_atual = request.args.get('pagina', 1, type=int)
    registros_por_pagina = request.args.get('registros_por_pagina', 50, type=int)
    if not (1 <= registros_por_pagina <= 100):
        registros_por_pagina = 50
        flash("Lote de registros inválido; usando o lote padrão de 50.", 'warning')
    
    filtro_status_id = request.args.get('status', type=int)
    filtro_tipo = request.args.get('tipo', type=int)
    filtro_busca = proteger_input(request.args.get('busca'))
    
    # --- Validação e formatação de datas de filtro ---
    filtro_data_inicio = request.args.get('data_inicio')
    if filtro_data_inicio:
        try:
            datetime.strptime(filtro_data_inicio, '%Y-%m-%d')
        except ValueError:
            flash("Formato de 'Data Início' inválido. Use AAAA-MM-DD.", 'danger')
            filtro_data_inicio = None

    filtro_data_fim = request.args.get('data_fim')
    if filtro_data_fim:
        try:
            datetime.strptime(filtro_data_fim, '%Y-%m-%d')
        except ValueError:
            flash("Formato de 'Data Fim' inválido. Use AAAA-MM-DD.", 'danger')
            filtro_data_fim = None
    
    if filtro_data_inicio and filtro_data_fim:
        if datetime.strptime(filtro_data_inicio, '%Y-%m-%d') > datetime.strptime(filtro_data_fim, '%Y-%m-%d'):
            flash("Data de início não pode ser posterior à data de fim.", 'danger')
            filtro_data_inicio = None
            filtro_data_fim = None
    # --- Fim Validação de Datas ---

    filtro_envolve_notas = request.args.get('envolve_notas')
    if filtro_envolve_notas is not None:
        try:
            filtro_envolve_notas = int(filtro_envolve_notas)
            if filtro_envolve_notas not in [0, 1]:
                filtro_envolve_notas = None
        except ValueError:
            filtro_envolve_notas = None
    ordenar = request.args.get('ordenar', 'prazo_asc')
    ordenar_opcoes = ['data_entrada_asc', 'data_entrada_desc', 'titular_asc', 'titular_desc', 'tipo_asc', 'tipo_desc', 'status_asc', 'status_desc', 'id_asc', 'id_desc', 'matricula_asc', 'matricula_desc', 'prazo_asc', 'prazo_desc']
    if ordenar not in ordenar_opcoes: ordenar = 'prazo_asc'
    
    status_pendente_obj = obter_status_processo_config()
    status_pendente_ids_to_filter = []
    excluded_status_names = ['Finalizado']
    for status_item in status_pendente_obj:
        if "Pendente" in status_item['nome'] and status_item['nome'] not in excluded_status_names:
            status_pendente_ids_to_filter.append(status_item['id'])

    filters = {
        'tipo': filtro_tipo, 'busca': filtro_busca,
        'data_inicio': filtro_data_inicio, 'data_fim': filtro_data_fim, 'envolve_notas': filtro_envolve_notas,
        'filtro_pendentes_dashboard': 1
    }
    
    if filtro_status_id is not None:
        filters['status_id'] = filtro_status_id

    try:
        resultado = listar_processos(filters, pagina_atual, registros_por_pagina, ordenar)
        processos = resultado['processos']
        total_registros = resultado['total_records']
        total_paginas = resultado['total_pages']
        tipos_servico = obter_tipos_servico()
        status_list = obter_status_processo_config()
        prazo_vencido = get_overdue_processes_count()
        processos_andamento = get_in_progress_processes_count()
        processos_criados_hoje = get_today_processes_count() 

        has_active_filters = any([
            (filtro_status_id is not None and filtro_status_id not in status_pendente_ids_to_filter) or \
            (filtro_status_id is None and status_pendente_ids_to_filter),
            filtro_tipo, filtro_busca, filtro_data_inicio, filtro_data_fim,
            filtro_envolve_notas is not None,
            ordenar != 'prazo_asc', registros_por_pagina != 50
        ])
        
        acao = request.args.get('acao')
        if acao == 'cadastrado': flash('Processo cadastrado com sucesso.', 'success')
        elif acao == 'atualizado': flash('Registro atualizado com sucesso.', 'success')
        elif acao == 'excluido': flash('Registro excluído com sucesso.', 'success')
        elif acao == 'exportado_excel': flash('Dados exportados para Excel com sucesso!', 'success')
        if not processos and has_active_filters:
            flash('Nenhum processo pendente encontrado com os filtros aplicados.', 'info')
    except Exception as e:
        logger.exception(f"Erro ao carregar processos pendentes: {e}")
        flash('Erro ao carregar processos. Por favor, tente novamente ou contate o suporte.', 'danger')
        return redirect(url_for('auth.dashboard'))
    return render_template('processos/pendentes.html',
                           processos=processos, total_registros=total_registros, total_paginas=total_paginas,
                           pagina_atual=pagina_atual, registros_por_pagina=registros_por_pagina,
                           filtro_status_id=filtro_status_id,
                           filtro_tipo=filtro_tipo, filtro_busca=filtro_busca,
                           filtro_data_inicio=filtro_data_inicio, filtro_data_fim=filtro_data_fim,
                           filtro_envolve_notas=filtro_envolve_notas, ordenar=ordenar,
                           has_active_filters=has_active_filters,
                           tipos_servico=tipos_servico,
                           status_list=status_list,
                           prazo_vencido=prazo_vencido, processos_andamento=processos_andamento,
                           processos_criados_hoje=processos_criados_hoje,
                           get_contrast_color=get_contrast_color, formatar_data=formatar_data)

@processos_bp.route('/vinculados', methods=['GET'])
@login_status_required
@permission_required('processos_visualizar')
def vinculados():
    logger.info(f"Acessando processos vinculados ao usuário logado. Usuário ID: {session.get('usuario_id')}, IP: {get_client_ip()}")
    usuario_id_logado = session.get('usuario_id')
    pagina_atual = request.args.get('pagina', 1, type=int)
    registros_por_pagina = request.args.get('registros_por_pagina', 50, type=int)
    ordenar = request.args.get('ordenar', 'id_desc')
    
    filtro_status_id = request.args.get('status', type=int)
    filtro_tipo = request.args.get('tipo', type=int)
    filtro_busca = proteger_input(request.args.get('busca'))
    filtro_data_inicio = request.args.get('data_inicio')
    filtro_data_fim = request.args.get('data_fim')
    filtro_envolve_notas = request.args.get('envolve_notas')

    # --- Validação e formatação de datas de filtro ---
    if filtro_data_inicio:
        try:
            datetime.strptime(filtro_data_inicio, '%Y-%m-%d')
        except ValueError:
            flash("Formato de 'Data Início' inválido. Use AAAA-MM-DD.", 'danger')
            filtro_data_inicio = None

    if filtro_data_fim:
        try:
            datetime.strptime(filtro_data_fim, '%Y-%m-%d')
        except ValueError:
            flash("Formato de 'Data Fim' inválido. Use AAAA-MM-DD.", 'danger')
            filtro_data_fim = None
    
    if filtro_data_inicio and filtro_data_fim:
        if datetime.strptime(filtro_data_inicio, '%Y-%m-%d') > datetime.strptime(filtro_data_fim, '%Y-%m-%d'):
            flash("Data de início não pode ser posterior à data de fim.", 'danger')
            filtro_data_inicio = None
            filtro_data_fim = None
    # --- Fim Validação de Datas ---

    if filtro_envolve_notas is not None:
        try:
            filtro_envolve_notas = int(filtro_envolve_notas)
            if filtro_envolve_notas not in [0, 1]:
                filtro_envolve_notas = None
        except ValueError:
            filtro_envolve_notas = None

    filters = {
        'responsavel_id': usuario_id_logado,
        'status_id': filtro_status_id,
        'tipo': filtro_tipo,
        'busca': filtro_busca,
        'data_inicio': filtro_data_inicio,
        'data_fim': filtro_data_fim,
        'envolve_notas': filtro_envolve_notas,
    }

    try:
        resultado = listar_processos(filters, pagina_atual, registros_por_pagina, ordenar)
        processos = resultado['processos']
        total_registros = resultado['total_records']
        total_paginas = resultado['total_pages']
        tipos_servico = obter_tipos_servico()
        status_list = obter_status_processo_config()
        prazo_vencido = get_overdue_processes_count()
        processos_andamento = get_in_progress_processes_count()
        processos_criados_hoje = get_today_processes_count() 

        has_active_filters = any([
            filtro_status_id, filtro_tipo, filtro_busca, filtro_data_inicio,
            filtro_data_fim, filtro_envolve_notas is not None,
            ordenar != 'id_desc',
            registros_por_pagina != 50
        ])

        acao = request.args.get('acao')
        if acao == 'cadastrado': flash('Processo cadastrado com sucesso.', 'success')
        elif acao == 'atualizado': flash('Registro atualizado com sucesso.', 'success')
        elif acao == 'excluido': flash('Registro excluído com sucesso.', 'success')
        elif acao == 'exportado_excel': flash('Dados exportados para Excel com sucesso!', 'success')
        if not processos and has_active_filters:
            flash('Nenhum processo vinculado encontrado com os filtros aplicados.', 'info')
    except Exception as e:
        logger.exception(f"Erro ao carregar processos vinculados: {e}")
        flash('Erro ao carregar processos. Por favor, tente novamente ou contate o suporte.', 'danger')
        return redirect(url_for('auth.dashboard'))
    return render_template('processos/vinculados.html',
                           processos=processos, total_registros=total_registros, total_paginas=total_paginas,
                           pagina_atual=pagina_atual, registros_por_pagina=registros_por_pagina,
                           filtro_status_id=filtro_status_id, filtro_tipo=filtro_tipo, filtro_busca=filtro_busca,
                           filtro_data_inicio=filtro_data_inicio, filtro_data_fim=filtro_data_fim,
                           filtro_envolve_notas=filtro_envolve_notas, ordenar=ordenar,
                           has_active_filters=has_active_filters,
                           tipos_servico=tipos_servico,
                           status_list=status_list,
                           prazo_vencido=prazo_vencido, processos_andamento=processos_andamento,
                           processos_criados_hoje=processos_criados_hoje,
                           get_contrast_color=get_contrast_color, formatar_data=formatar_data)


@processos_bp.route('/em_andamento', methods=['GET'])
@login_status_required
@permission_required('processos_visualizar')
def em_andamento():
    logger.info(f"Acessando processos em andamento. Usuário ID: {session.get('usuario_id')}, IP: {get_client_ip()}")
    pagina_atual = request.args.get('pagina', 1, type=int)
    registros_por_pagina = request.args.get('registros_por_pagina', 50, type=int)
    if not (1 <= registros_por_pagina <= 100):
        registros_por_pagina = 50
        flash("Lote de registros inválido; usando o lote padrão de 50.", 'warning')

    filtro_status_id = request.args.get('status', type=int)
    filtro_tipo = request.args.get('tipo', type=int)
    filtro_busca = proteger_input(request.args.get('busca'))

    filtro_data_inicio = request.args.get('data_inicio')
    if filtro_data_inicio:
        try:
            datetime.strptime(filtro_data_inicio, '%Y-%m-%d')
        except ValueError:
            flash("Formato de 'Data Início' inválido. Use AAAA-MM-DD.", 'danger')
            filtro_data_inicio = None

    filtro_data_fim = request.args.get('data_fim')
    if filtro_data_fim:
        try:
            datetime.strptime(filtro_data_fim, '%Y-%m-%d')
        except ValueError:
            flash("Formato de 'Data Fim' inválido. Use AAAA-MM-DD.", 'danger')
            filtro_data_fim = None

    if filtro_data_inicio and filtro_data_fim:
        if datetime.strptime(filtro_data_inicio, '%Y-%m-%d') > datetime.strptime(filtro_data_fim, '%Y-%m-%d'):
            flash("Data de início não pode ser posterior à data de fim.", 'danger')
            filtro_data_inicio = None
            filtro_data_fim = None

    filtro_envolve_notas = request.args.get('envolve_notas')
    if filtro_envolve_notas is not None:
        try:
            filtro_envolve_notas = int(filtro_envolve_notas)
            if filtro_envolve_notas not in [0, 1]:
                filtro_envolve_notas = None
        except ValueError:
            filtro_envolve_notas = None

    ordenar = request.args.get('ordenar', 'id_desc')
    ordenar_opcoes = ['data_entrada_asc', 'data_entrada_desc', 'titular_asc', 'titular_desc',
                      'tipo_asc', 'tipo_desc', 'status_asc', 'status_desc',
                      'id_asc', 'id_desc', 'matricula_asc', 'matricula_desc', 'prazo_asc', 'prazo_desc']
    if ordenar not in ordenar_opcoes:
        ordenar = 'id_desc'

    filters = {
        'tipo': filtro_tipo,
        'busca': filtro_busca,
        'data_inicio': filtro_data_inicio,
        'data_fim': filtro_data_fim,
        'envolve_notas': filtro_envolve_notas,
        'filtro_em_andamento': 1
    }
    if filtro_status_id is not None:
        filters['status_id'] = filtro_status_id

    try:
        resultado = listar_processos(filters, pagina_atual, registros_por_pagina, ordenar)
        processos = resultado['processos']
        total_registros = resultado['total_records']
        total_paginas = resultado['total_pages']
        tipos_servico = obter_tipos_servico()
        status_list = obter_status_processo_config()
        prazo_vencido = get_overdue_processes_count()
        processos_andamento = get_em_andamento_processes_count()
        processos_criados_hoje = get_today_processes_count()

        has_active_filters = any([
            filtro_status_id, filtro_tipo, filtro_busca,
            filtro_data_inicio, filtro_data_fim,
            filtro_envolve_notas is not None,
            ordenar != 'id_desc', registros_por_pagina != 50
        ])

        acao = request.args.get('acao')
        if acao == 'cadastrado': flash('Processo cadastrado com sucesso.', 'success')
        elif acao == 'atualizado': flash('Registro atualizado com sucesso.', 'success')
        elif acao == 'excluido': flash('Registro excluído com sucesso.', 'success')
        elif acao == 'exportado_excel': flash('Dados exportados para Excel com sucesso!', 'success')
        if not processos and has_active_filters:
            flash('Nenhum processo em andamento encontrado com os filtros aplicados.', 'info')
    except Exception as e:
        logger.exception(f"Erro ao carregar processos em andamento: {e}")
        flash('Erro ao carregar processos. Por favor, tente novamente ou contate o suporte.', 'danger')
        return redirect(url_for('auth.dashboard'))

    return render_template('processos/em_andamento.html',
                           processos=processos, total_registros=total_registros,
                           total_paginas=total_paginas, pagina_atual=pagina_atual,
                           registros_por_pagina=registros_por_pagina,
                           filtro_status_id=filtro_status_id, filtro_tipo=filtro_tipo,
                           filtro_busca=filtro_busca, filtro_data_inicio=filtro_data_inicio,
                           filtro_data_fim=filtro_data_fim, filtro_envolve_notas=filtro_envolve_notas,
                           ordenar=ordenar, has_active_filters=has_active_filters,
                           tipos_servico=tipos_servico, status_list=status_list,
                           prazo_vencido=prazo_vencido, processos_andamento=processos_andamento,
                           processos_criados_hoje=processos_criados_hoje,
                           get_contrast_color=get_contrast_color, formatar_data=formatar_data)


@processos_bp.route('/visualizar/processo=<int:processo_id>')
@login_status_required
@permission_required('processos_visualizar')
def visualizar(processo_id):
    processo_detalhes = get_processo_by_id(processo_id)
    if not processo_detalhes:
        flash("Processo não encontrado.", 'danger')
        return redirect(url_for('processos.todos'))
    
    anexos_do_processo_raw = obter_anexos_processo(processo_id)
    anexos_do_processo = []
    for anexo_raw in anexos_do_processo_raw:
        anexo = dict(anexo_raw)
        anexo_path = os.path.join(get_upload_folder(), secure_filename(anexo['nome_arquivo']))
        anexo['existe_no_servidor'] = os.path.exists(anexo_path) and os.path.isfile(anexo_path)
        anexos_do_processo.append(anexo)

    historico_alteracoes = obter_historico_processo(processo_id)

    csrf_token_val = gerar_csrf_token()

    return render_template('processos/visualizar.html',
                           processo=processo_detalhes,
                           anexos_do_processo=anexos_do_processo,
                           historico_alteracoes=historico_alteracoes,
                           get_contrast_color=get_contrast_color,
                           formatar_data=formatar_data,
                           get_icone_anexo=obter_icone_anexo,
                           formatar_tamanho_arquivo=formatar_tamanho_arquivo,
                           csrf_token=csrf_token_val)

@processos_bp.route('/editar/<int:processo_id>', methods=['GET', 'POST'])
@login_status_required
@permission_required('processos_editar')
def editar(processo_id):
    usuario_id = session.get('usuario_id')
    usuario_role = session.get('usuario_role')
    is_ajax_request = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    tipos_processo = [t for t in obter_tipos_servico() if t.get('ativo')]
    status_processo = [s for s in obter_status_processo_config() if s.get('ativo')]
    responsaveis = obter_usuarios_para_selecao()

    dados_form = {}
    anexos_existentes = []

    if request.method == 'GET':
        processo_original = get_processo_by_id(processo_id)
        if not processo_original:
            if is_ajax_request: return jsonify(success=False, message="Processo não encontrado.", redirect=url_for('processos.todos'), type='danger'), 404
            else: flash("Processo não encontrado.", 'danger'); return redirect(url_for('processos.todos'))

        # Processo finalizado: permite edição mas sinaliza para o template exibir aviso
        processo_finalizado = (
            processo_original.get('status_nome_original') == 'Finalizado' or
            bool(processo_original.get('data_conclusao'))
        )

        lock_result = acquire_lock('processos', processo_id, usuario_id, LOCK_TIMEOUT_MINUTES)
        if lock_result is not True:
            if is_ajax_request: return jsonify(success=False, message=lock_result.get('error', "Falha ao adquirir bloqueio de edição."), type='danger'), 409
            else: flash(lock_result.get('error', "Outro usuário está editando este processo ou o bloqueio não pôde ser adquirido."), 'danger'); return redirect(url_for('processos.visualizar', processo_id=processo_id))

        try:
            anexos_existentes_brutos = obter_anexos_processo(processo_id)
            for anexo_raw in anexos_existentes_brutos:
                anexo = dict(anexo_raw)
                anexo_path = os.path.join(get_upload_folder(), secure_filename(anexo['nome_arquivo']))
                anexo['existe_no_servidor'] = os.path.exists(anexo_path) and os.path.isfile(anexo_path)
                anexos_existentes.append(anexo)
        except Exception as e:
            release_lock('processos', processo_id, usuario_id)
            logger.exception(f"Erro ao carregar anexos do processo {processo_id} para edição: {e}")
            if is_ajax_request: return jsonify(success=False, message="Erro ao carregar anexos para edição.", type='danger'), 500
            else: flash("Erro ao carregar anexos para edição.", 'danger'); return redirect(url_for('processos.todos'))

        logger.info(f"Acessando formulário de edição para processo ID: {processo_id}. Usuário ID: {usuario_id}")
        
        dados_form = {
            'titular': processo_original['titular'],
            'titular_id': processo_original.get('titular_id') or '',
            'titular_telefone': processo_original.get('titular_telefone') or '',
            'titular_email': processo_original.get('titular_email') or '',
            'tipo_servico': processo_original['tipo_id'],
            'matricula': processo_original['matricula'],
            'possui_matricula': processo_original.get('possui_matricula') or 0,
            'status': processo_original['status_nome_original'],
            'apresentante': processo_original['apresentante'],
            'apresentante_id': processo_original.get('apresentante_id') or '',
            'apresentante_telefone': processo_original['apresentante_telefone'],
            'apresentante_email': processo_original['apresentante_email'],
            'prazo_final': processo_original['prazo_final'],
            'envolvido_notas': processo_original['envolvido_notas'], 'observacoes': processo_original['observacoes'],
            'data_entrada': processo_original['data_entrada'],
            'updated_at': processo_original.get('updated_at') or ''
        }
        
        if dados_form['data_entrada']:
            try:
                dt_obj_entrada = datetime.strptime(dados_form['data_entrada'], '%Y-%m-%d %H:%M:%S')
            except ValueError:
                dt_obj_entrada = datetime.strptime(dados_form['data_entrada'], '%Y-%m-%d')
            dados_form['data_entrada'] = dt_obj_entrada.strftime('%Y-%m-%d')
        else:
            dados_form['data_entrada'] = ''
            
        if dados_form['prazo_final']:
            try:
                dt_obj_prazo = datetime.strptime(dados_form['prazo_final'], '%Y-%m-%d %H:%M:%S')
            except ValueError:
                dt_obj_prazo = datetime.strptime(dados_form['prazo_final'], '%Y-%m-%d')
            dados_form['prazo_final'] = dt_obj_prazo.strftime('%Y-%m-%d')
        else:
            dados_form['prazo_final'] = ''


        csrf_token_val = gerar_csrf_token()
        return render_template('processos/editar.html',
                            processo=processo_original,
                            processo_id=processo_id,
                            dados_form=dados_form,
                            tipos_processo=tipos_processo,
                            status_processo=status_processo,
                            responsaveis=responsaveis,
                            anexos_existentes=anexos_existentes,
                            get_icone_anexo=obter_icone_anexo,
                            formatar_tamanho_arquivo=formatar_tamanho_arquivo,
                            get_contrast_color=get_contrast_color,
                            formatar_data=formatar_data,
                            csrf_token=csrf_token_val,
                            LOCK_TIMEOUT_MINUTES=LOCK_TIMEOUT_MINUTES,
                            processo_finalizado=processo_finalizado)

    elif request.method == 'POST':
        dados_form.update({
            'titular': request.form.get('titular', ''),
            'titular_id': request.form.get('titular_id', ''),
            'titular_telefone': request.form.get('titular_telefone', ''),
            'titular_email': request.form.get('titular_email', ''),
            'tipo_servico': request.form.get('tipo_servico', type=int),
            'matricula': request.form.get('matricula', ''),
            'possui_matricula': '1' if request.form.get('possui_matricula') == '1' else '0',
            'status': request.form.get('status', ''),
            'apresentante': request.form.get('apresentante', ''),
            'apresentante_id': request.form.get('apresentante_id', ''),
            'apresentante_telefone': request.form.get('apresentante_telefone', ''),
            'apresentante_email': request.form.get('apresentante_email', ''),
            'prazo_final': request.form.get('prazo_final', ''),
            'envolvido_notas': request.form.get('envolvido_notas', type=int),
            'observacoes': request.form.get('observacoes', ''),
            'data_entrada': request.form.get('data_entrada', ''),
            'updated_at': request.form.get('updated_at', '')
        })

        processo_original = get_processo_by_id(processo_id)
        if not processo_original:
            return jsonify(success=False, message="Processo não encontrado para atualização.", type='danger'), 404

        expected_updated_at = (dados_form.get('updated_at') or '').strip() or None
        if expected_updated_at and processo_original.get('updated_at') != expected_updated_at:
            logger.warning(
                f"Conflito de versão ao editar processo {processo_id}. "
                f"Esperado: {expected_updated_at}; atual: {processo_original.get('updated_at')}"
            )
            return jsonify(
                success=False,
                message="Este processo foi alterado por outro usuário. Recarregue a tela antes de salvar.",
                type='warning',
                conflict=True,
            ), 409

        # Processo finalizado: permite salvar mas registra aviso no log
        if processo_original.get('status_nome_original') == 'Finalizado' or processo_original.get('data_conclusao'):
            logger.warning(f"Usuário ID {usuario_id} editou processo finalizado ID {processo_id}.")
        
        # O bloco de permissão foi comentado para permitir que qualquer utilizador
        # autenticado possa editar um processo, conforme solicitado.
        '''
        if usuario_role != 'admin' and processo_original['responsavel_id'] != usuario_id:
            flash("Você não tem permissão para editar este processo.", 'error')
            logger.warning(f"Tentativa de edição (POST) não autorizada por role/responsável. Processo ID: {processo_id}. Usuário: {usuario_id} (Role: {usuario_role}). Responsável: {processo_original['responsavel_id']}")
            return jsonify(success=False, message="Você não tem permissão para editar este processo.", type='danger'), 403
        '''

        if not verificar_csrf_token(request.form.get('csrf_token')): 
            return jsonify(success=False, message="Token de segurança inválido. Por favor, recarregue a página e tente novamente.", type='danger'), 403
        
        lock_result = acquire_lock('processos', processo_id, usuario_id, LOCK_TIMEOUT_MINUTES)
        if lock_result is not True:
            return jsonify(success=False, message=lock_result.get('error', "Falha ao adquirir/renovar bloqueio de edição. Por favor, recarregue a página."), type='danger'), 409
        
        titular = proteger_input(dados_form['titular'])
        titular_cadastro_id = request.form.get('titular_id', type=int) or processo_original.get('titular_id')
        titular_telefone = proteger_input(dados_form['titular_telefone'])
        titular_email = proteger_input(dados_form['titular_email'])
        tipo_id = dados_form['tipo_servico']
        possui_matricula = 1 if request.form.get('possui_matricula') == '1' else 0
        matricula = proteger_input(dados_form['matricula']) if possui_matricula else None
        status_nome_post = proteger_input(dados_form['status'])
        apresentante = proteger_input(dados_form['apresentante'])
        apresentante_cadastro_id = request.form.get('apresentante_id', type=int) or processo_original.get('apresentante_id')
        apresentante_telefone = proteger_input(dados_form['apresentante_telefone'])
        apresentante_email = proteger_input(dados_form['apresentante_email'])
        prazo_final_input = dados_form['prazo_final']
        envolvido_notas = dados_form['envolvido_notas']
        observacoes = proteger_input(dados_form['observacoes'])
        data_entrada_do_form = dados_form['data_entrada']
        excluir_anexos_ids = request.form.getlist('excluir_anexos[]')
        
        campos_obrigatorios = {
            'titular': 'Titular', 'tipo_servico': 'Tipo de Serviço',
            'status': 'Status',
            'data_entrada': 'Data de Entrada'
        }
        campos_faltantes = [nome for campo, nome in campos_obrigatorios.items() if not dados_form.get(campo) and dados_form.get(campo) != 0]
        if campos_faltantes:
            msg = f"Os seguintes campos são obrigatórios: {', '.join(campos_faltantes)}"
            release_lock('processos', processo_id, usuario_id)
            if not is_ajax_request:
                flash(msg, 'danger')
                return render_template('processos/editar.html',
                                       processo=processo_original,
                                       processo_id=processo_id,
                                       dados_form=dados_form,
                                       tipos_processo=tipos_processo,
                                       status_processo=status_processo,
                                       responsaveis=responsaveis,
                                       anexos_existentes=obter_anexos_processo(processo_id),
                                       get_icone_anexo=obter_icone_anexo,
                                       formatar_tamanho_arquivo=formatar_tamanho_arquivo,
                                       get_contrast_color=get_contrast_color,
                                       formatar_data=formatar_data,
                                       csrf_token=gerar_csrf_token(),
                                       LOCK_TIMEOUT_MINUTES=LOCK_TIMEOUT_MINUTES)
            return jsonify(success=False, message=msg, type='danger', field_error='form_fields_missing'), 400
        
        try:
            validar_tipo_servico(tipo_id)
            validar_status(status_nome_post)
            from models import validar_telefone_unico, validar_email_unico
            if matricula:
                validar_formato_matricula(matricula, processo_id)
            if apresentante_telefone: 
                validar_telefone(apresentante_telefone)
                validar_telefone_unico(apresentante_telefone, processo_id=processo_id, titular_nome=titular)
            if apresentante_email: 
                validar_email(apresentante_email)
                validar_email_unico(apresentante_email, processo_id=processo_id, titular_nome=titular)
            
            if not data_entrada_do_form:
                raise ValueError("A Data de Entrada é obrigatória.")
            try:
                data_entrada_obj = datetime.strptime(data_entrada_do_form, '%Y-%m-%d')
            except ValueError:
                raise ValueError("Data de Entrada inválida. Use o formato AAAA-MM-DD.")

            prazo_final_calculado = None
            if prazo_final_input:
                try:
                    data_prazo_obj = datetime.strptime(prazo_final_input, '%Y-%m-%d')
                    prazo_final_calculado = data_prazo_obj.strftime('%Y-%m-%d')
                except ValueError as e:
                    raise ValueError(f"Formato de Prazo Final inválido. Use AAAA-MM-DD. Erro: {e}")
            else:
                tipo_servico_info = executar_query("SELECT prazo_padrao FROM tipos_servico WHERE id = ?", [tipo_id], fetch_one=True)
                dias_prazo = tipo_servico_info['prazo_padrao'] if tipo_servico_info else 30
                prazo_final_calculado = (data_entrada_obj + timedelta(days=dias_prazo)).strftime('%Y-%m-%d')
                logger.info(f"Prazo final calculado automaticamente para {prazo_final_calculado} (Data Entrada: {data_entrada_do_form}, Prazo Padrão: {dias_prazo} dias) durante a edição.")

        except ValueError as e:
            field_error_map = {
                "Telefone inválido": "apresentante_telefone", "E-mail inválido": "apresentante_email",
                "MATRÍCULA:": "matricula", "TELEFONE:": "apresentante_telefone", "E-MAIL:": "apresentante_email",
                "Tipo de serviço com ID": "tipo_servico",
                "Status": "status",
                "Formato de Prazo Final inválido.": "prazo_final", "O Prazo Final é obrigatório.": "prazo_final",
                "A Data de Entrada é obrigatória.": "data_entrada", "Data de Entrada inválida.": "data_entrada",
                "Matrícula inválida": "matricula"
            }
            field = next((v for k, v in field_error_map.items() if k in str(e)), None)
            release_lock('processos', processo_id, usuario_id)
            if not is_ajax_request:
                flash(str(e), 'danger')
                return render_template('processos/editar.html',
                                       processo=processo_original,
                                       processo_id=processo_id,
                                       dados_form=dados_form,
                                       tipos_processo=tipos_processo,
                                       status_processo=status_processo,
                                       responsaveis=responsaveis,
                                       anexos_existentes=obter_anexos_processo(processo_id),
                                       get_icone_anexo=obter_icone_anexo,
                                       formatar_tamanho_arquivo=formatar_tamanho_arquivo,
                                       get_contrast_color=get_contrast_color,
                                       formatar_data=formatar_data,
                                       csrf_token=gerar_csrf_token(),
                                       LOCK_TIMEOUT_MINUTES=LOCK_TIMEOUT_MINUTES)
            return jsonify(success=False, message=str(e), type='danger', field_error=field), 400
        
        status_id_para_update = executar_query("SELECT id FROM status_processo WHERE nome = ?", [status_nome_post], fetch_one=True)['id']
        data_conclusao_update = processo_original['data_conclusao']

        if status_nome_post.lower() in ['finalizado'] and \
           processo_original['status_nome_original'].lower() not in ['finalizado']:
            data_conclusao_update = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        elif status_nome_post.lower() not in ['finalizado'] and \
             processo_original['status_nome_original'].lower() in ['finalizado']:
            data_conclusao_update = None

        try:
            with get_sqlite_connection() as conn:
                anexos_excluidos_nomes, erros_exclusao_anexos = excluir_anexos_selecionados(excluir_anexos_ids, processo_id, usuario_id, conn)
                if erros_exclusao_anexos: logger.warning(f"Erros ao excluir anexos durante edição (DB/físico): {'; '.join(erros_exclusao_anexos)}")
                
                anexos_salvos_novos, arquivos_rejeitados = processar_anexos_upload(request.files, processo_id, usuario_id, conn)
                
                # Atualiza o cadastro existente quando o nome foi alterado; só cria
                # outro cadastro quando o nome anterior é compartilhado por outros processos.
                titular_id = upsert_titular_from_processo(
                    titular,
                    titular_telefone,
                    titular_email,
                    processo_id,
                    connection=conn,
                    nome_anterior=processo_original.get('titular'),
                    cadastro_id=titular_cadastro_id,
                    usuario_id=usuario_id,
                    processo_excluido_id=processo_id,
                )
                apresentante_id = upsert_apresentante_from_processo(
                    apresentante,
                    apresentante_telefone,
                    apresentante_email,
                    processo_id,
                    connection=conn,
                    nome_anterior=processo_original.get('apresentante'),
                    cadastro_id=apresentante_cadastro_id,
                    usuario_id=usuario_id,
                    processo_excluido_id=processo_id,
                )
                
                rows_affected = update_processo(
                    processo_id=processo_id,
                    titular=titular,
                    titular_telefone=titular_telefone,
                    titular_email=titular_email,
                    matricula=matricula,
                    tipo_id=tipo_id,
                    data_entrada=data_entrada_do_form,
                    status_id=status_id_para_update,
                    prazo_final=prazo_final_calculado,
                    apresentante=apresentante,
                    apresentante_telefone=apresentante_telefone,
                    apresentante_email=apresentante_email,
                    responsavel_id=usuario_id,
                    envolvido_notas=envolvido_notas,
                    observacoes=observacoes,
                    data_conclusao=data_conclusao_update,
                    possui_matricula=possui_matricula,
                    connection=conn,
                    titular_id=titular_id,
                    apresentante_id=apresentante_id,
                    expected_updated_at=expected_updated_at
                )
                
                if not rows_affected:
                    logger.info(f"Nenhuma alteração detectada no registro principal do processo {processo_id}.")

                # Monta o contexto com TODAS as alterações realizadas (exibido no modal de detalhe)
                _linhas_contexto = []

                # Status
                _old_status = (processo_original.get('status_nome_original') or '').strip()
                if _old_status.lower() != status_nome_post.lower():
                    _linhas_contexto.append(f"Status: '{_old_status}' → '{status_nome_post}'")
                    if status_nome_post.lower() in ['finalizado']:
                        _linhas_contexto.append("Processo marcado como finalizado.")

                # Tipo de serviço
                _novo_tipo_info = executar_query(
                    "SELECT nome FROM tipos_servico WHERE id = ?", [tipo_id], fetch_one=True, connection=conn
                )
                _novo_tipo_nome = _novo_tipo_info['nome'] if _novo_tipo_info else str(tipo_id)
                _old_tipo_nome = (processo_original.get('tipo_nome') or '').strip()
                if _old_tipo_nome.lower() != _novo_tipo_nome.lower():
                    _linhas_contexto.append(f"Tipo de serviço: '{_old_tipo_nome}' → '{_novo_tipo_nome}'")

                # Titular
                _old_titular = (processo_original.get('titular') or '').strip()
                _new_titular = (titular or '').strip()
                if _old_titular.lower() != _new_titular.lower():
                    _linhas_contexto.append(f"Titular: '{_old_titular}' → '{_new_titular}'")

                # Matrícula
                _old_matricula = (processo_original.get('matricula') or '').strip()
                _new_matricula = (matricula or '').strip()
                if _old_matricula != _new_matricula:
                    _linhas_contexto.append(f"Matrícula: '{_old_matricula}' → '{_new_matricula}'")

                # Data de Entrada
                _old_data_entrada = (processo_original.get('data_entrada') or '').strip()[:10]
                _new_data_entrada = (data_entrada_do_form or '').strip()[:10]
                if _old_data_entrada != _new_data_entrada:
                    def _fmt_data(d):
                        try:
                            from datetime import datetime as _dt
                            return _dt.strptime(d, '%Y-%m-%d').strftime('%d/%m/%Y')
                        except Exception:
                            return d or '—'
                    _linhas_contexto.append(f"Data de entrada: '{_fmt_data(_old_data_entrada)}' → '{_fmt_data(_new_data_entrada)}'")

                # Prazo Final
                _old_prazo = (processo_original.get('prazo_final') or '').strip()[:10]
                _new_prazo = (prazo_final_calculado or '').strip()[:10]
                if _old_prazo != _new_prazo:
                    def _fmt_prazo(d):
                        try:
                            from datetime import datetime as _dt
                            return _dt.strptime(d, '%Y-%m-%d').strftime('%d/%m/%Y')
                        except Exception:
                            return d or '—'
                    _linhas_contexto.append(f"Prazo final: '{_fmt_prazo(_old_prazo)}' → '{_fmt_prazo(_new_prazo)}'")

                # Telefone do apresentante
                _old_tel = (processo_original.get('apresentante_telefone') or '').strip()
                _new_tel = (apresentante_telefone or '').strip()
                if _old_tel != _new_tel:
                    _linhas_contexto.append(f"Telefone do apresentante: '{_old_tel}' → '{_new_tel}'")

                # E-mail do apresentante
                _old_email = (processo_original.get('apresentante_email') or '').strip()
                _new_email = (apresentante_email or '').strip()
                if _old_email != _new_email:
                    _linhas_contexto.append(f"E-mail do apresentante: '{_old_email}' → '{_new_email}'")

                # Apresentante
                _old_apres = (processo_original.get('apresentante') or '').strip()
                _new_apres = (apresentante or '').strip()
                if _old_apres.lower() != _new_apres.lower():
                    _linhas_contexto.append(f"Apresentante: '{_old_apres}' → '{_new_apres}'")

                # Observações
                _old_obs = (processo_original.get('observacoes') or '').strip()
                _new_obs = (observacoes or '').strip()
                if _old_obs != _new_obs:
                    _old_obs_val = 'Sim' if _old_obs else 'Não'
                    _new_obs_val = 'Sim' if _new_obs else 'Não'
                    _linhas_contexto.append(f"Observações: '{_old_obs_val}' → '{_new_obs_val}'")

                # Envolve notas
                _old_notas = processo_original.get('envolvido_notas') or 0
                _new_notas = envolvido_notas or 0
                if int(_old_notas) != int(_new_notas):
                    _linhas_contexto.append(f"Envolve notas: '{'Sim' if _old_notas else 'Não'}' → '{'Sim' if _new_notas else 'Não'}'")

                # Possui anexos (informativo — sempre incluído)
                _total_anexos = executar_query(
                    "SELECT COUNT(*) AS c FROM anexos_processos WHERE processo_id = ?",
                    [processo_id], fetch_one=True, connection=conn
                )
                _qtd_anexos = _total_anexos['c'] if _total_anexos else 0
                _linhas_contexto.append(f"[anexos] {'Sim' if _qtd_anexos > 0 else 'Não'} ({_qtd_anexos} arquivo{'s' if _qtd_anexos != 1 else ''})")

                _contexto_edit = "\n".join(_linhas_contexto) if _linhas_contexto else None

                logger.info(f"Processo '{titular}' (ID: {processo_id}) atualizado.")

            # Logs gravados APÓS o commit (conexão própria)
            gravar_log(
                "Editou processo",
                processo_id, usuario_id, get_client_ip(),
                descricao=f"Processo: {titular} (Matrícula: {matricula})",
                contexto=_contexto_edit
            )

            # A libertação do bloqueio ocorre aqui, após o bloco 'with' ser finalizado,
            # para evitar o erro "database is locked" no SQLite.
            release_lock('processos', processo_id, usuario_id)

            # Registra logs individuais de anexo APÓS o commit da transação principal
            # (conexão própria para não interferir na transação)
            for _nome_anx in anexos_salvos_novos:
                gravar_log('Adicionou anexo', processo_id, usuario_id, get_client_ip(), f"Anexo: {_nome_anx}")
            
            msg_sucesso = "Processo atualizado com sucesso!"
            if anexos_salvos_novos: msg_sucesso += f" {len(anexos_salvos_novos)} novo(s) anexo(s) adicionado(s)."
            if arquivos_rejeitados: msg_sucesso += f" Atenção: {len(arquivos_rejeitados)} arquivo(s) não foram anexados. Detalhes: {', '.join(arquivos_rejeitados)}."
            if erros_exclusao_anexos: msg_sucesso += f" Atenção: Erro na exclusão de {len(erros_exclusao_anexos)} anexo(s) existente(s)."
            
            return jsonify(success=True, message=msg_sucesso, type='success', redirect=url_for('processos.visualizar', processo_id=processo_id))
        
        except sqlite3.Error as e:
            release_lock('processos', processo_id, usuario_id)
            error_msg = f"Erro de banco de dados ao atualizar processo: {str(e)}"
            logger.error(error_msg, exc_info=True)
            field_error_map = {
                "unique constraint failed": "matricula"
            }
            field = next((v for k, v in field_error_map.items() if k in str(e)), None)

            if not is_ajax_request:
                flash(error_msg, 'danger')
                return render_template('processos/editar.html',
                                       processo=processo_original,
                                       processo_id=processo_id,
                                       dados_form=dados_form,
                                       tipos_processo=tipos_processo,
                                       status_processo=status_processo,
                                       responsaveis=responsaveis,
                                       anexos_existentes=obter_anexos_processo(processo_id),
                                       get_icone_anexo=obter_icone_anexo,
                                       formatar_tamanho_arquivo=formatar_tamanho_arquivo,
                                       get_contrast_color=get_contrast_color,
                                       formatar_data=formatar_data,
                                       csrf_token=gerar_csrf_token(),
                                       LOCK_TIMEOUT_MINUTES=LOCK_TIMEOUT_MINUTES)
            return jsonify(success=False, message=error_msg, type='danger', field_error=field), 500
        except ValueError as e:
            release_lock('processos', processo_id, usuario_id)
            error_msg = str(e)
            logger.warning(f"Erro de validação ou lógica ao atualizar processo: {error_msg}", exc_info=True)
            if not is_ajax_request:
                flash(error_msg, 'info')
                return render_template('processos/editar.html',
                                       processo=processo_original,
                                       processo_id=processo_id,
                                       dados_form=dados_form,
                                       tipos_processo=tipos_processo,
                                       status_processo=status_processo,
                                       responsaveis=responsaveis,
                                       anexos_existentes=obter_anexos_processo(processo_id),
                                       get_icone_anexo=obter_icone_anexo,
                                       formatar_tamanho_arquivo=formatar_tamanho_arquivo,
                                       get_contrast_color=get_contrast_color,
                                       formatar_data=formatar_data,
                                       csrf_token=gerar_csrf_token(),
                                       LOCK_TIMEOUT_MINUTES=LOCK_TIMEOUT_MINUTES)
            return jsonify(success=False, message=error_msg, type='info'), 400
        except Exception as e:
            release_lock('processos', processo_id, usuario_id)
            error_msg = f"Erro inesperado ao atualizar processo: {str(e)}"
            logger.error(error_msg, exc_info=True)
            if not is_ajax_request:
                flash(error_msg, 'danger')
                return render_template('processos/editar.html',
                                       processo=processo_original,
                                       processo_id=processo_id,
                                       dados_form=dados_form,
                                       tipos_processo=tipos_processo,
                                       status_processo=status_processo,
                                       responsaveis=responsaveis,
                                       anexos_existentes=obter_anexos_processo(processo_id),
                                       get_icone_anexo=obter_icone_anexo,
                                       formatar_tamanho_arquivo=formatar_tamanho_arquivo,
                                       get_contrast_color=get_contrast_color,
                                       formatar_data=formatar_data,
                                       csrf_token=gerar_csrf_token(),
                                       LOCK_TIMEOUT_MINUTES=LOCK_TIMEOUT_MINUTES)
            return jsonify(success=False, message=error_msg, type='danger'), 500


@processos_bp.route('/download_anexo/<int:anexo_id>')
@login_status_required
@permission_required('processos_anexos')
def download_anexo(anexo_id):
    usuario_id = session.get('usuario_id')
    logger.info(f"Tentativa de download de anexo ID: {anexo_id} por Usuário ID: {usuario_id}, IP: {get_client_ip()}")
    anexo_info = executar_query("SELECT processo_id, nome_original, nome_arquivo, tipo FROM anexos_processos WHERE id = ? LIMIT 1", [anexo_id], fetch_one=True)
    
    if not anexo_info:
        flash("Anexo não encontrado.", 'danger'); logger.warning(f"Tentativa de download de anexo inexistente: ID {anexo_id}")
        return redirect(url_for('processos.todos'))
    
    filename_on_server = secure_filename(anexo_info['nome_arquivo'])
    file_path = os.path.join(get_upload_folder(), filename_on_server)
    
    if not os.path.exists(file_path) or not os.path.isfile(file_path) or \
       not os.path.abspath(file_path).startswith(os.path.abspath(get_upload_folder())):
        flash("Arquivo não encontrado no servidor ou acesso negado.", 'danger')
        logger.error(f"Arquivo físico não encontrado ou path traversal detectado para anexo ID {anexo_id}: {file_path}")
        return redirect(url_for('processos.todos'))
    
    try:
        gravar_log("Baixou anexo", anexo_info['processo_id'], usuario_id, get_client_ip(), f"Anexo: {anexo_info['nome_original']}", connection=None)
        return send_from_directory(get_upload_folder(), filename_on_server, as_attachment=True, download_name=anexo_info['nome_original'], mimetype=anexo_info['tipo'])
    except Exception as e:
        logger.error(f"Erro ao servir arquivo para download '{filename_on_server}': {e}", exc_info=True)
        flash("Erro ao baixar o anexo. Tente novamente.", 'danger')
        return redirect(url_for('processos.todos'))

@processos_bp.route('/excluir', methods=['POST'])
@login_status_required
@permission_required('processos_excluir')
def excluir():
    processo_id = request.form.get('id', type=int)
    usuario_id = session.get('usuario_id')
    
    if not verificar_csrf_token(request.form.get('csrf_token')):
        flash("Token de segurança inválido.", 'danger')
        return redirect(url_for('processos.todos'))

    if not processo_id:
        flash("ID do processo para exclusão não fornecido.", 'danger')
        return redirect(url_for('processos.todos'))

    # --- INÍCIO DA ATUALIZAÇÃO ---
    # Verifica se o processo está sendo editado por outro usuário antes de prosseguir.
    lock_info = is_record_locked('processos', processo_id, usuario_id)
    if lock_info:
        user_locking = lock_info.get('user_nome', 'outro usuário')
        flash(f"Não foi possível excluir o processo. Ele está sendo editado por: {user_locking}.", 'danger')
        logger.warning(f"Tentativa de exclusão do processo {processo_id} bloqueada, pois está em edição por {user_locking} (ID: {lock_info.get('user_id')}).")
        return redirect(url_for('processos.visualizar', processo_id=processo_id))
    # --- FIM DA ATUALIZAÇÃO ---

    # Obtém os detalhes ANTES da transação para usar nas mensagens
    processo_info = get_processo_by_id(processo_id)
    if not processo_info:
        flash(f"Processo ID {processo_id} não encontrado.", 'danger')
        return redirect(url_for('processos.todos'))

    nome_processo_log = f"#{processo_info['id']} ('{processo_info['titular']}')"
    
    # Apaga os arquivos físicos primeiro
    anexos_do_processo = obter_anexos_processo(processo_id)
    for anexo in anexos_do_processo:
        try:
            caminho_arquivo = os.path.join(get_upload_folder(), secure_filename(anexo['nome_arquivo']))
            if os.path.exists(caminho_arquivo):
                os.remove(caminho_arquivo)
                logger.info(f"Anexo físico '{anexo['nome_arquivo']}' removido do servidor.")
        except Exception as e:
            logger.error(f"Falha ao remover arquivo físico '{anexo['nome_arquivo']}': {e}", exc_info=True)
            flash(f"Erro ao apagar o anexo físico '{anexo['nome_original']}'. A exclusão foi cancelada.", 'danger')
            return redirect(url_for('processos.todos'))

    try:
        # O 'with' garante uma transação atómica: ou tudo funciona, ou nada é alterado.
        with get_sqlite_connection() as conn:
            
            # 1. Chama a função de exclusão do modelo, passando a conexão partilhada
            excluir_processo_db(processo_id, connection=conn)
            
            # 2. Grava o log da operação, usando a MESMA conexão
            # NOTA: processo_id=None porque o processo já foi deletado (FK CASCADE causaria erro)
            gravar_log(
                acao="Exclusão de Processo",
                processo_id=None,
                usuario_id=usuario_id,
                ip=get_client_ip(),
                descricao=f"Processo {nome_processo_log} foi permanentemente excluído.",
                connection=conn
            )
        
        # Se o bloco 'with' for finalizado, a transação é confirmada (commit).
        flash(f"Processo {nome_processo_log} excluído com sucesso!", 'success')

    except Exception as e:
        # Se ocorrer um erro dentro do 'with', a transação é revertida (rollback).
        logger.error(f"Falha na transação ao excluir o processo {processo_id}: {e}", exc_info=True)
        flash(f"Ocorreu um erro ao excluir o processo. A operação foi cancelada. Erro: {e}", 'danger')

    return redirect(url_for('processos.todos'))

# Em routes/processos.py (pode ser antes da rota de imprimir_lista)

@processos_bp.route('/exportar_excel')
@login_status_required
@permission_required('processos_exportar')
def exportar_excel():
    logger.info(f"Usuário {session.get('usuario_id')} solicitou a exportação de processos para Excel. IP: {get_client_ip()}")

    # 1. Obter os mesmos filtros que o usuário está vendo na tela
    ordenar = request.args.get('ordenar', 'id_desc')
    filters = {
        'status_id': request.args.get('status', type=int),
        'tipo': request.args.get('tipo', type=int),
        'busca': proteger_input(request.args.get('busca')),
        'data_inicio': request.args.get('data_inicio'),
        'data_fim': request.args.get('data_fim'),
        'envolve_notas': request.args.get('envolve_notas', type=int),
        'responsavel_id': request.args.get('responsavel_id', type=int),
        'filtro_pendentes_dashboard': request.args.get('filtro_pendentes_dashboard', type=int)
    }

    try:
        # 2. Buscar TODOS os processos que correspondem aos filtros (sem paginação)
        resultado = listar_processos(filters, 1, 99999, ordenar)
        processos = resultado['processos']

        if not processos:
            flash('Nenhum processo encontrado para exportar com os filtros atuais.', 'info')
            return redirect(request.referrer or url_for('processos.todos'))

        # 3. Criar o arquivo Excel em memória
        wb = Workbook()
        ws = wb.active
        ws.title = "Processos"

        # 4. Adicionar e estilizar os cabeçalhos
        headers = ["ID", "Titular", "Tipo de Serviço", "Matrícula", "Data de Entrada", "Status", "Prazo Final", "Responsável"]
        ws.append(headers)
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="8B6332", end_color="8B6332", fill_type="solid") # Dourado Escuro da sua paleta

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # 5. Preencher as linhas com os dados dos processos
        for processo in processos:
            # Preparar dados para a linha
            data_entrada_str = formatar_data(processo.get('data_entrada'))
            prazo_final_str = formatar_data(processo.get('prazo_final'))
            
            row_data = [
                processo.get('id'),
                processo.get('titular'),
                processo.get('tipo_nome'),
                processo.get('matricula'),
                data_entrada_str,
                processo.get('status_nome'),
                prazo_final_str,
                processo.get('responsavel_nome')
            ]
            ws.append(row_data)

            # Estilizar a célula de Status com a cor correspondente
            hex_color = processo.get('status_hex', 'FFFFFF').lstrip('#')
            font_color = get_contrast_color(f"#{hex_color}").lstrip('#')
            
            status_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
            status_font = Font(color=font_color)

            # A coluna de Status é a 6ª (F)
            status_cell = ws.cell(row=ws.max_row, column=6)
            status_cell.fill = status_fill
            status_cell.font = status_font

        # 6. Ajustar a largura das colunas
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 25
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 30

        # 7. Salvar o arquivo em um buffer de memória e enviá-lo como resposta
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"Relatorio_Processos_{timestamp}.xlsx"

        gravar_log(f"Exportou {len(processos)} processos para Excel", None, session.get('usuario_id'), get_client_ip())

        return Response(output,
                        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        headers={"Content-Disposition": f"attachment;filename={filename}"})

    except Exception as e:
        logger.exception(f"Erro ao exportar processos para Excel: {e}")
        flash(f"Ocorreu um erro ao gerar o arquivo Excel: {e}", 'danger')
        return redirect(request.referrer or url_for('processos.todos'))

# --- Nova Rota para Impressão da Lista de Processos ---
def _build_lista_filters_and_title(view_mode):
    """
    Constrói o dicionário de filtros e o título descritivo para a rota imprimir_lista
    e gerar_pdf_lista, respeitando os filtros implícitos de cada visão de lista.
    """
    view_mode_labels = {
        'todos':        'Todos os Processos',
        'em_andamento': 'Processos em Andamento',
        'pendentes':    'Processos Pendentes',
        'vinculados':   'Meus Processos',
        'hoje':         'Processos de Hoje',
    }
    titulo_relatorio = view_mode_labels.get(view_mode, 'Processos')

    status_filter = request.args.get('status', '').strip()
    filtro_status_id = None
    if status_filter:
        try:
            filtro_status_id = int(status_filter)
        except ValueError:
            pass

    status_ids_in_str = request.args.get('status_ids_in', '').strip()
    status_ids_in = []
    if status_ids_in_str:
        try:
            status_ids_in = [int(s_id) for s_id in status_ids_in_str.split(',') if s_id.isdigit()]
        except ValueError:
            pass

    filtro_tipo        = request.args.get('tipo', type=int)
    filtro_busca       = proteger_input(request.args.get('busca'))
    filtro_data_inicio = request.args.get('data_inicio')
    filtro_data_fim    = request.args.get('data_fim')
    filtro_envolve_notas_str = request.args.get('envolve_notas')
    filtro_envolve_notas = None
    if filtro_envolve_notas_str is not None:
        try:
            filtro_envolve_notas = int(filtro_envolve_notas_str)
            if filtro_envolve_notas not in [0, 1]:
                filtro_envolve_notas = None
        except (ValueError, TypeError):
            filtro_envolve_notas = None
    ordenar = request.args.get('ordenar', 'id_desc')

    # Filtros base (manuais do usuário)
    filters = {
        'tipo':           filtro_tipo,
        'busca':          filtro_busca,
        'data_inicio':    filtro_data_inicio,
        'data_fim':       filtro_data_fim,
        'envolve_notas':  filtro_envolve_notas,
    }

    # Aplica filtros implícitos conforme a visão de origem
    if view_mode == 'em_andamento':
        filters['filtro_em_andamento'] = 1
        if filtro_status_id is not None:
            filters['status_id'] = filtro_status_id
    elif view_mode == 'pendentes':
        filters['filtro_pendentes_dashboard'] = 1
        if filtro_status_id is not None:
            filters['status_id'] = filtro_status_id
    elif view_mode == 'vinculados':
        # O responsavel_id vem da sessão — não é exposto na URL
        filters['responsavel_id'] = session.get('usuario_id')
        if filtro_status_id is not None:
            filters['status_id'] = filtro_status_id
    elif view_mode == 'hoje':
        hoje_fmt = datetime.now().strftime('%Y-%m-%d')
        filters['data_inicio'] = filtro_data_inicio or hoje_fmt
        filters['data_fim']    = filtro_data_fim    or hoje_fmt
        if filtro_status_id is not None:
            filters['status_id'] = filtro_status_id
    else:  # 'todos' ou qualquer outro
        if status_ids_in:
            filters['status_ids_in'] = status_ids_in
        else:
            filters['status_id'] = filtro_status_id

    tipos_servico = obter_tipos_servico() or []
    status_list = obter_status_processo_config() or []
    tipo_label = next((row['nome'] for row in tipos_servico if int(row['id']) == filtro_tipo), None) if filtro_tipo else None
    status_ids_context = status_ids_in or ([filtro_status_id] if filtro_status_id else [])
    status_labels = [row['nome'] for row in status_list if int(row['id']) in status_ids_context]
    contexto_filtros = []
    if filtro_busca:
        contexto_filtros.append(f'Busca: {filtro_busca}')
    if status_labels:
        contexto_filtros.append(f"Status: {', '.join(status_labels)}")
    elif status_ids_context:
        contexto_filtros.append(f'Status: {", ".join(str(item) for item in status_ids_context)}')
    if tipo_label:
        contexto_filtros.append(f'Tipo: {tipo_label}')
    if filters.get('data_inicio') or filters.get('data_fim'):
        inicio = filters.get('data_inicio') or 'início'
        fim = filters.get('data_fim') or 'fim'
        try:
            inicio = datetime.strptime(inicio, '%Y-%m-%d').strftime('%d/%m/%Y') if inicio != 'início' else inicio
            fim = datetime.strptime(fim, '%Y-%m-%d').strftime('%d/%m/%Y') if fim != 'fim' else fim
        except (TypeError, ValueError):
            pass
        contexto_filtros.append(f'Período: {inicio} a {fim}')
    if filtro_envolve_notas is not None:
        contexto_filtros.append(f'Envolve notas: {"Sim" if filtro_envolve_notas else "Não"}')
    if not contexto_filtros:
        contexto_filtros.append('Sem filtros adicionais')
    contexto_relatorio = {
        'filtros': contexto_filtros,
        'ordenacao': ordenar,
        'visao': titulo_relatorio,
    }
    return filters, ordenar, titulo_relatorio, contexto_relatorio


def _resumir_prazos_relatorio(processos):
    """Resume a situação de prazo somente dos processos ativos do relatório."""
    hoje = datetime.now().date()
    vencidos = 0
    em_dia = 0
    sem_prazo = 0
    for processo in processos:
        prazo = str(processo.get('prazo_final') or '')[:10]
        if processo.get('data_conclusao'):
            continue
        if not prazo:
            sem_prazo += 1
            continue
        try:
            if datetime.strptime(prazo, '%Y-%m-%d').date() < hoje:
                vencidos += 1
            else:
                em_dia += 1
        except ValueError:
            sem_prazo += 1
    return {'vencidos': vencidos, 'em_dia': em_dia, 'sem_prazo': sem_prazo}


@processos_bp.route('/imprimir_lista', methods=['GET'])
@login_status_required
@permission_required('processos_imprimir')
def imprimir_lista():
    logger.info(f"Gerando relatório de impressão da lista de processos para Usuário ID: {session.get('usuario_id')}, IP: {get_client_ip()}")

    view_mode = request.args.get('view_mode', 'todos')
    filters, ordenar, titulo_relatorio, contexto_relatorio = _build_lista_filters_and_title(view_mode)

    try:
        resultado = listar_processos(filters, 1, 9999, ordenar)
        processos = resultado['processos']
        total_registros_impressao = resultado['total_records']
        prazo_resumo = _resumir_prazos_relatorio(processos)

        empresa_info = get_empresa_info()
        logo_filename = empresa_info.get('logo') if empresa_info else None
        logo_url = get_image_url_for_display(logo_filename, is_company_logo=True)

        gravar_log("Imprimiu lista de processos", None, session.get('usuario_id'), get_client_ip(),
                   f"view_mode={view_mode}, Filtros: {request.args.to_dict()}, Registros: {total_registros_impressao}", connection=None)

        return render_template('relatorios/processos_print.html',
                               processos=processos,
                               total_registros=total_registros_impressao,
                               titulo_relatorio=titulo_relatorio,
                               contexto_relatorio=contexto_relatorio,
                               prazo_resumo=prazo_resumo,
                               logo_url=logo_url,
                               now=datetime.now(),
                               formatar_data=formatar_data,
                               get_contrast_color=get_contrast_color)

    except Exception as e:
        logger.exception(f"Erro ao gerar lista de impressão de processos: {e}")
        flash('Erro ao gerar relatório de impressão. Por favor, tente novamente.', 'danger')
        return redirect(url_for('processos.todos'))


@processos_bp.route('/gerar_pdf_lista', methods=['GET'])
@login_status_required
@permission_required('processos_imprimir')
def gerar_pdf_lista():
    """Gera PDF da listagem de processos respeitando todos os filtros ativos."""
    view_mode = request.args.get('view_mode', 'todos')
    filters, ordenar, titulo_relatorio, contexto_relatorio = _build_lista_filters_and_title(view_mode)

    try:
        resultado = listar_processos(filters, 1, 9999, ordenar)
        processos = resultado['processos']
        total_registros_impressao = resultado['total_records']
        prazo_resumo = _resumir_prazos_relatorio(processos)

        empresa_info = get_empresa_info()
        logo_filename = empresa_info.get('logo') if empresa_info else None
        logo_url = get_image_url_for_display(logo_filename, is_company_logo=True, for_pdf=True)

        html_string = render_template(
            'relatorios/processos_print.html',
            processos=processos,
            total_registros=total_registros_impressao,
            titulo_relatorio=titulo_relatorio,
            contexto_relatorio=contexto_relatorio,
            prazo_resumo=prazo_resumo,
            logo_url=logo_url,
            now=datetime.now(),
            formatar_data=formatar_data,
            get_contrast_color=get_contrast_color
        )

        pdf_bytes = HTML(string=html_string, base_url=request.url_root).write_pdf()

        filename = f"Lista_Processos_{view_mode}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"

        gravar_log("Baixou PDF de lista de processos", None, session.get('usuario_id'), get_client_ip(),
                   f"view_mode={view_mode}, Registros: {total_registros_impressao}", connection=None)

        return Response(
            pdf_bytes,
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )

    except Exception as e:
        logger.exception(f"Erro ao gerar PDF de lista de processos: {e}")
        flash('Erro ao gerar PDF da listagem. Por favor, tente novamente.', 'danger')
        return redirect(url_for(f'processos.{view_mode}' if view_mode in ['todos', 'em_andamento', 'pendentes', 'vinculados', 'hoje'] else 'processos.todos'))


def _build_relatorio_processo_context(processo, for_pdf=False):
    """Monta o contexto único usado pela impressão e pelo PDF detalhado."""
    empresa_info = get_empresa_info()
    logo_filename = empresa_info.get('logo') if empresa_info else None

    return {
        'processo': processo,
        'historico': obter_historico_processo(processo['id']),
        'formatar_data': formatar_data,
        'formatar_tamanho_arquivo': formatar_tamanho_arquivo,
        'get_icone_anexo': obter_icone_anexo,
        'get_contrast_color': get_contrast_color,
        'now': datetime.now(),
        'logo_url': get_image_url_for_display(
            logo_filename,
            is_company_logo=True,
            for_pdf=for_pdf,
        ),
    }


def _render_relatorio_processo_pdf(processo, template_context=None):
    """Renderiza o mesmo relatório detalhado usado na impressão em PDF."""
    context = template_context or _build_relatorio_processo_context(processo, for_pdf=True)
    html_string = render_template('relatorios/relatorio_processo_print.html', **context)
    pdf_bytes = HTML(string=html_string, base_url=request.url_root).write_pdf()
    filename = f"Relatorio_Processo_{processo['id']}.pdf"

    return Response(
        pdf_bytes,
        mimetype='application/pdf',
        headers={'Content-Disposition': f'attachment;filename={filename}'},
    )


# --- Novas Rotas para Geração de Relatórios Personalizados ---

@processos_bp.route('/gerar_relatorio_customizado/<int:processo_id>/<string:tipo>')
@login_status_required
@permission_required('processos_relatorio')
def gerar_relatorio_customizado(processo_id, tipo):
    processo = get_processo_by_id(processo_id)

    if not processo:
        flash('Processo não encontrado para gerar relatório.', 'error')
        return redirect(url_for('processos.todos'))

    try:
        if tipo == 'html_print':
            context = _build_relatorio_processo_context(processo, for_pdf=False)
            return render_template('relatorios/relatorio_processo_print.html', **context)

        if tipo == 'pdf':
            context = _build_relatorio_processo_context(processo, for_pdf=True)
            return _render_relatorio_processo_pdf(processo, context)

        flash('Tipo de relatório inválido.', 'error')
        return redirect(url_for('processos.visualizar', processo_id=processo_id))
    except Exception as e:
        logger.error(
            f"Erro ao gerar relatório do processo {processo_id} ({tipo}): {e}",
            exc_info=True,
        )
        flash('Ocorreu um erro ao gerar o relatório do processo.', 'danger')
        return redirect(url_for('processos.visualizar', processo_id=processo_id))

@processos_bp.route('/historico_print/<int:processo_id>')
@login_status_required
@permission_required('processos_visualizar')
def imprimir_historico(processo_id):
    """Gera página de impressão exclusiva do histórico de alterações do processo."""
    processo = get_processo_by_id(processo_id)
    if not processo:
        flash('Processo não encontrado.', 'danger')
        return redirect(url_for('processos.todos'))

    historico = obter_historico_processo(processo_id)

    empresa_info = get_empresa_info()
    logo_filename = empresa_info.get('logo') if empresa_info else None
    logo_url = get_image_url_for_display(logo_filename, is_company_logo=True)

    gravar_log("Imprimiu histórico de alterações", processo_id, session.get('usuario_id'), get_client_ip(),
               f"Processo: {processo.get('titular')} (ID: {processo_id})", connection=None)

    return render_template('relatorios/historico_processo_print.html',
                           processo=processo,
                           historico=historico,
                           logo_url=logo_url,
                           now=datetime.now(),
                           formatar_data=formatar_data)


@processos_bp.route('/release-lock-ajax', methods=['POST'])
@login_status_required
def release_lock_ajax():
    """
    Endpoint para liberar o bloqueio de um registro via AJAX/Beacon,
    tipicamente chamado pelo evento 'beforeunload' no frontend.
    """
    try:
        # O método sendBeacon com JSON envia os dados no corpo do pedido.
        data = request.get_json(silent=True)
        if not data:
            logger.warning(f"Chamada a release-lock-ajax sem corpo JSON. IP: {get_client_ip()}")
            return jsonify({'success': False, 'error': 'Request sem dados'}), 400

        csrf_token = (
            request.headers.get('X-CSRFToken')
            or request.headers.get('X-CSRF-Token')
            or data.get('csrf_token')
        )
        if not verificar_csrf_token(csrf_token):
            logger.warning(f"CSRF inválido ao liberar lock via AJAX. IP: {get_client_ip()}")
            return jsonify({'success': False, 'error': 'Token de segurança inválido.'}), 403

        record_id = data.get('record_id')
        table_name = data.get('table_name')
        usuario_id = session.get('usuario_id')

        if table_name not in ALLOWED_LOCK_TABLES:
            return jsonify({'success': False, 'error': 'Tabela não permitida para bloqueio.'}), 400

        if record_id and table_name and usuario_id:
            # A função release_lock (do models.py) lida com a lógica de libertação.
            release_lock(table_name, int(record_id), usuario_id)
            # Retorna uma resposta vazia com status 204 (No Content), ideal para beacons.
            return '', 204
        else:
            logger.warning(f"Chamada a release-lock-ajax com dados ausentes: {data}. IP: {get_client_ip()}")
            return jsonify({'success': False, 'error': 'Dados em falta no request'}), 400

    except Exception as e:
        logger.error(f"Erro ao libertar lock via AJAX/Beacon: {e}", exc_info=True)
        return jsonify({'success': False, 'error': 'Não foi possível concluir a liberação do registro agora. Tente novamente.'}), 500
        
@processos_bp.route('/gerar_pdf/<int:processo_id>')
@login_status_required
@permission_required('processos_pdf')
def gerar_pdf(processo_id):
    try:
        processo = get_processo_by_id(processo_id)
        if not processo:
            flash('Processo não encontrado.', 'danger')
            return redirect(url_for('processos.todos'))

        context = _build_relatorio_processo_context(processo, for_pdf=True)
        return _render_relatorio_processo_pdf(processo, context)
    except Exception as e:
        logger.error(f"Erro ao gerar PDF para processo {processo_id}: {e}", exc_info=True)
        flash('Ocorreu um erro ao gerar o relatório em PDF.', 'danger')
        return redirect(url_for('processos.visualizar', processo_id=processo_id))