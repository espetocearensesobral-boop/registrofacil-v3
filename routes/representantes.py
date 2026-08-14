
from flask import Blueprint, render_template, request, jsonify, session, flash, redirect, url_for
from models import (
    listar_representantes, get_representante_by_id, get_historico_servicos_representante, 
    buscar_representantes_json, executar_query, get_empresa_info, gravar_log,
    representante_tem_processos, editar_representante, excluir_representante
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from io import BytesIO
from flask import Response
from datetime import datetime
from routes.auth import login_status_required, get_client_ip, proteger_input
from routes.permissoes import permission_required
from utils.logger import logger

representantes_bp = Blueprint('representantes', __name__, url_prefix='/representantes')

@representantes_bp.route('/', methods=['GET'])
@login_status_required
@permission_required('representantes_visualizar')
def index():
    logger.info(f"Acessando lista de representantes. Usuário ID: {session.get('usuario_id')}, IP: {get_client_ip()}")
    
    pagina_atual = request.args.get('pagina', 1, type=int)
    registros_por_pagina = request.args.get('registros_por_pagina', 10, type=int)
    busca = proteger_input(request.args.get('busca', ''))
    ordenar = request.args.get('ordenar', 'nome')
    direcao = request.args.get('direcao', 'asc')
    
    filtros = {'busca': busca, 'ordenar': ordenar, 'direcao': direcao}
    
    try:
        resultado = listar_representantes(filtros, pagina_atual, registros_por_pagina)
        representantes = resultado['representantes']
        total_registros = resultado['total_records']
        total_paginas = resultado['total_pages']
        
        return render_template('representantes/index.html',
                               representantes=representantes,
                               total_registros=total_registros,
                               total_paginas=total_paginas,
                               pagina_atual=pagina_atual,
                               registros_por_pagina=registros_por_pagina,
                               busca=busca,
                               ordenar=ordenar,
                               direcao=direcao)
    except Exception as e:
        logger.error(f"Erro ao listar representantes: {e}", exc_info=True)
        flash("Erro ao carregar lista de representantes.", "danger")
        return redirect(url_for('auth.dashboard'))

@representantes_bp.route('/editar/<int:representante_id>', methods=['GET', 'POST'])
@login_status_required
@permission_required('representantes_editar')
def editar(representante_id):
    representante = get_representante_by_id(representante_id)
    if not representante:
        flash("Representante não encontrado.", "warning")
        return redirect(url_for('representantes.index'))

    if request.method == 'POST':
        nome = proteger_input(request.form.get('nome', '').strip().upper())
        telefone = proteger_input(request.form.get('telefone', '').strip())
        email = proteger_input(request.form.get('email', '').strip().lower())

        if not nome:
            flash("O nome do representante é obrigatório.", "danger")
            return render_template('representantes/editar.html', representante=representante)

        try:
            from utils.helpers import validar_telefone, validar_email
            if telefone: validar_telefone(telefone)
            if email: validar_email(email)

            existente = executar_query(
                "SELECT id FROM representantes WHERE nome = ? AND id != ?", [nome, representante_id], fetch_one=True
            )
            if existente:
                flash(f"Já existe outro representante com o nome '{nome}'.", "warning")
                return render_template('representantes/editar.html', representante=representante)

            editar_representante(representante_id, nome, telefone if telefone else None, email if email else None)
            gravar_log("Editou representante", None, session.get('usuario_id'), get_client_ip(), f"Representante ID {representante_id}: {nome}")
            flash("Representante atualizado com sucesso!", "success")
            return redirect(url_for('representantes.visualizar', representante_id=representante_id))

        except ValueError as e:
            flash(str(e), "danger")
            return render_template('representantes/editar.html', representante=representante)
        except Exception as e:
            logger.error(f"Erro ao editar representante {representante_id}: {e}", exc_info=True)
            flash("Erro ao atualizar representante.", "danger")

    return render_template('representantes/editar.html', representante=representante)

@representantes_bp.route('/excluir/<int:representante_id>', methods=['POST'])
@login_status_required
@permission_required('representantes_excluir')
def excluir(representante_id):
    from routes.auth import verificar_csrf_token
    if not verificar_csrf_token(request.form.get('csrf_token')):
        flash("Token de segurança inválido.", "danger")
        return redirect(url_for('representantes.index'))

    representante = get_representante_by_id(representante_id)
    if not representante:
        flash("Representante não encontrado.", "warning")
        return redirect(url_for('representantes.index'))

    if representante_tem_processos(representante_id):
        flash("Não é possível excluir este representante pois existem processos vinculados a ele.", "danger")
        return redirect(url_for('representantes.index'))

    try:
        excluir_representante(representante_id)
        gravar_log("Excluiu representante", None, session.get('usuario_id'), get_client_ip(), f"Representante excluído: {representante['nome']}")
        flash("Representante excluído com sucesso!", "success")
    except Exception as e:
        logger.error(f"Erro ao excluir representante {representante_id}: {e}", exc_info=True)
        flash("Erro ao excluir representante.", "danger")

    return redirect(url_for('representantes.index'))

@representantes_bp.route('/visualizar/<int:representante_id>', methods=['GET'])
@login_status_required
@permission_required('representantes_visualizar')
def visualizar(representante_id):
    representante = get_representante_by_id(representante_id)
    if not representante:
        flash("Representante não encontrado.", "warning")
        return redirect(url_for('representantes.index'))

    historico = get_historico_servicos_representante(representante['nome'])
    return render_template('representantes/visualizar.html', representante=representante, historico=historico)

@representantes_bp.route('/novo', methods=['GET', 'POST'])
@login_status_required
@permission_required('representantes_criar')
def novo():
    if request.method == 'POST':
        nome = proteger_input(request.form.get('nome', '').strip().upper())
        telefone = proteger_input(request.form.get('telefone', '').strip())
        email = proteger_input(request.form.get('email', '').strip().lower())

        if not nome:
            flash("O nome do representante é obrigatório.", "danger")
            return render_template('representantes/novo.html')

        try:
            from utils.helpers import validar_telefone, validar_email
            if telefone: validar_telefone(telefone)
            if email: validar_email(email)

            existente = executar_query("SELECT id FROM representantes WHERE nome = ?", [nome], fetch_one=True)
            if existente:
                flash(f"Já existe um representante cadastrado com o nome '{nome}'.", "warning")
                return render_template('representantes/novo.html')

            now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            with executar_query("INSERT INTO representantes (nome, telefone, email, created_at, updated_at) VALUES (?, ?, ?, ?, ?)", 
                                [nome, telefone if telefone else None, email if email else None, now, now], fetch_one=False) as conn:
                pass

            gravar_log("Criou representante", None, session.get('usuario_id'), get_client_ip(), f"Novo representante: {nome}")
            flash("Representante cadastrado com sucesso!", "success")
            return redirect(url_for('representantes.index'))

        except ValueError as e:
            flash(str(e), "danger")
            return render_template('representantes/novo.html')
        except Exception as e:
            logger.error(f"Erro ao criar representante: {e}", exc_info=True)
            flash("Erro ao cadastrar representante.", "danger")

    return render_template('representantes/novo.html')

@representantes_bp.route('/exportar', methods=['GET'])
@login_status_required
@permission_required('representantes_exportar')
def exportar():
    try:
        resultado = listar_representantes(registros_por_pagina=1000000)
        representantes = resultado['representantes']

        wb = Workbook()
        ws = wb.active
        ws.title = "Representantes"

        headers = ['ID', 'Nome', 'Telefone', 'E-mail', 'Total de Processos', 'Data de Cadastro']
        ws.append(headers)

        header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font

        for rep in representantes:
            ws.append([
                rep['id'],
                rep['nome'],
                rep['telefone'] or '',
                rep['email'] or '',
                rep['total_processos'],
                rep['created_at']
            ])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"representantes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment;filename={filename}"}
        )
    except Exception as e:
        logger.error(f"Erro ao exportar representantes: {e}", exc_info=True)
        flash("Erro ao exportar representantes.", "danger")
        return redirect(url_for('representantes.index'))

@representantes_bp.route('/imprimir', methods=['GET'])
@login_status_required
@permission_required('representantes_imprimir')
def imprimir():
    """Renderiza a listagem completa de representantes para impressão."""
    busca = proteger_input(request.args.get('busca', ''))
    ordenar = request.args.get('ordenar', 'nome')
    direcao = request.args.get('direcao', 'asc')
    filtros = {'busca': busca, 'ordenar': ordenar, 'direcao': direcao}

    try:
        resultado = listar_representantes(filtros, 1, 9999)
        representantes = resultado['representantes']

        empresa_info = get_empresa_info()
        logo_filename = empresa_info.get('logo') if empresa_info else None
        from utils.file_uploads import get_image_url_for_display
        logo_url = get_image_url_for_display(logo_filename, is_company_logo=True)

        return render_template(
            'relatorios/representantes_print.html',
            representantes=representantes,
            total_registros=resultado['total_records'],
            logo_url=logo_url,
            busca=busca,
            now=datetime.now()
        )
    except Exception as e:
        logger.error(f"Erro ao gerar impressão de representantes: {e}", exc_info=True)
        flash("Erro ao gerar relatório de impressão.", "danger")
        return redirect(url_for('representantes.index'))

@representantes_bp.route('/api/buscar', methods=['GET'])
@login_status_required
def api_buscar():
    termo = request.args.get('q', '')
    representantes = buscar_representantes_json(termo)
    return jsonify(representantes)

@representantes_bp.route('/api/verificar-duplicidade', methods=['GET'])
@login_status_required
@permission_required('representantes_visualizar')
def api_verificar_duplicidade():
    """Verifica em tempo real se telefone ou e-mail já estão cadastrados para representantes."""
    campo = request.args.get('campo')
    valor = request.args.get('valor', '').strip()
    excluir_id = request.args.get('excluir_id', type=int)

    if campo not in ('telefone', 'email') or not valor:
        return jsonify({'duplicado': False})

    try:
        if campo == 'telefone':
            query = "SELECT id, nome FROM representantes WHERE telefone = ?"
        else:
            query = "SELECT id, nome FROM representantes WHERE email = ?"

        params = [valor]
        if excluir_id:
            query += " AND id != ?"
            params.append(excluir_id)

        resultado = executar_query(query, params, fetch_one=True)
        if resultado:
            return jsonify({'duplicado': True, 'representante': resultado['nome']})
        return jsonify({'duplicado': False})
    except Exception as e:
        logger.error(f"Erro ao verificar duplicidade de representante ({campo}): {e}")
        return jsonify({'duplicado': False})
